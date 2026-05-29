"""
==============================================================
  backend_logic.py  —  BACKEND (Logic Layer)
==============================================================

  [역할]
  "내부에서 돌아가는 코딩" — 사용자에게 보이지 않는 연산 담당.
  화면 출력이 단 한 줄도 없습니다.
  UI(Frontend)는 이 파일의 함수를 호출하기만 합니다.

  [포함 기능]
  ① 기상 데이터 수집 (API / WAMIS 크롤링 / 시뮬레이션)
  ② 토양 데이터 생성 (물 수지 모델)
  ③ 데이터 전처리 · 품질 관리
  ④ 관수 예측 모델 개발 · K-Fold 검증
  ⑤ 관수 의사결정 (모델 기반)
  ⑥ AI 권고문 생성 (Claude API / 규칙 기반)
  ⑦ CSV · Excel · JSON 저장 · 불러오기

  [Input / Output 규약]
  - 모든 public 함수는 순수 Python 자료형(dict, list, str, float)만 반환
  - 터미널 출력(print) 금지 — 로그는 Return 값에 포함
  - 예외는 모두 try/except로 포착하여 {"error": "..."} 형태로 반환
==============================================================
"""

import csv
import json
import math
import os
import random
import datetime
import statistics
from collections import Counter

# ── 선택적 패키지 (없어도 실행되도록 guard) ──────────────
try:
    import requests
    from bs4 import BeautifulSoup
    _REQUESTS_OK = True
except ImportError:
    _REQUESTS_OK = False

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    _EXCEL_OK = True
except ImportError:
    _EXCEL_OK = False

try:
    import pandas as pd
    _PANDAS_OK = True
except ImportError:
    _PANDAS_OK = False


# ══════════════════════════════════════════════════════════
#  섹션 A — 상수 · 도메인 지식
# ══════════════════════════════════════════════════════════

CROP_THRESHOLDS: dict = {
    "벼":            {"Min": 60, "Opt": 80, "Max": 100},
    "밀/보리":       {"Min": 40, "Opt": 60, "Max": 75},
    "옥수수":        {"Min": 45, "Opt": 65, "Max": 80},
    "콩":            {"Min": 40, "Opt": 58, "Max": 70},
    "감자":          {"Min": 50, "Opt": 70, "Max": 85},
    "토마토":        {"Min": 55, "Opt": 72, "Max": 85},
    "상추/채소":     {"Min": 50, "Opt": 68, "Max": 80},
    "과수(사과/배)": {"Min": 40, "Opt": 60, "Max": 75},
}

CROP_KC: dict = {
    "벼": 1.20, "밀/보리": 0.85, "옥수수": 1.15, "콩": 1.05,
    "감자": 1.10, "토마토": 1.15, "상추/채소": 1.00, "과수(사과/배)": 0.90,
}

SEASON_PARAMS: dict = {
    "봄":   {"Temp_Base": 15, "Temp_Var": 8,  "Humid_Base": 60, "Rain_Prob": 0.30},
    "여름": {"Temp_Base": 30, "Temp_Var": 5,  "Humid_Base": 78, "Rain_Prob": 0.45},
    "가을": {"Temp_Base": 18, "Temp_Var": 7,  "Humid_Base": 55, "Rain_Prob": 0.25},
    "겨울": {"Temp_Base": 2,  "Temp_Var": 8,  "Humid_Base": 45, "Rain_Prob": 0.20},
}

VALID_RANGE: dict = {
    "평균온도_C":      (-40.0, 50.0),
    "상대습도_Pct":    (0.0, 100.0),
    "강수량_mm":       (0.0, 300.0),
    "풍속_ms":         (0.0, 50.0),
    "일사량_MJ":       (0.0, 35.0),
    "증발산량ET0_mm":  (0.0, 15.0),
}

IRRIGATION_LABELS: dict = {0: "즉시관수", 1: "관수권장", 2: "적정상태", 3: "과습주의"}
FEATURE_COLS = [
    "토양수분_Pct", "평균온도_C", "상대습도_Pct",
    "강수량_mm", "증발산량ET0_mm", "풍속_ms",
    "최근3일강수_mm", "이동평균수분_Pct",
]


# ══════════════════════════════════════════════════════════
#  섹션 B — 데이터 수집
# ══════════════════════════════════════════════════════════

def _calc_et0(Mean_Temp: float, Temp_Var: float, Solar_MJ: float) -> float:
    """ET₀ 계산 (Penman-Monteith 간이식). 내부 전용."""
    try:
        return round(max(0.0, 0.0023 * (Mean_Temp + 17.8) * math.sqrt(max(0, Temp_Var)) * Solar_MJ * 0.1), 2)
    except Exception:
        return 0.5


def _sim_weather(Days: int, Season: str, Start_Date: datetime.date) -> list[dict]:
    """시뮬레이션 날씨 생성. 내부 전용."""
    P = SEASON_PARAMS.get(Season, SEASON_PARAMS["여름"])
    Result = []
    for I in range(Days):
        D     = Start_Date + datetime.timedelta(days=I)
        MT    = round(P["Temp_Base"] + random.uniform(-P["Temp_Var"], P["Temp_Var"]), 1)
        Rain  = round(random.uniform(3, 35), 1) if random.random() < P["Rain_Prob"] else 0.0
        Solar = round(random.uniform(8, 24), 1)
        Result.append({
            "날짜":            str(D),
            "출처":            "시뮬레이션",
            "최고온도_C":      round(MT + random.uniform(2, 5), 1),
            "최저온도_C":      round(MT - random.uniform(2, 5), 1),
            "평균온도_C":      MT,
            "상대습도_Pct":    round(min(100, max(10, P["Humid_Base"] + random.uniform(-20, 20))), 1),
            "강수량_mm":       Rain,
            "풍속_ms":         round(random.uniform(0.5, 7), 1),
            "일사량_MJ":       Solar,
            "증발산량ET0_mm":  _calc_et0(MT, P["Temp_Var"], Solar),
            "데이터품질_Flag": "S",
        })
    return Result


def collect_weather(
    Days: int = 30,
    Season: str = "여름",
    KMA_API_Key: str | None = None,
    Save_Dir: str = "data/raw",
) -> dict:
    """
    [INPUT]  Days, Season, KMA_API_Key, Save_Dir
    [OUTPUT] {"data": [...], "source": str, "saved_path": str, "log": str}

    기상청 API → WAMIS 크롤링 → 시뮬레이션 순으로 시도.
    데이터 출처:
      - 기상청 기상자료개방포털 https://data.kma.go.kr
      - WAMIS 한국수자원공사   https://www.wamis.go.kr
    """
    Today      = datetime.date.today()
    Start_Date = Today - datetime.timedelta(days=Days - 1)
    Log_Lines  = []
    Data: list[dict] | None = None
    Source = "시뮬레이션"

    # ◆ 기상청 API 시도
    if KMA_API_Key and _REQUESTS_OK:
        try:
            Params = {
                "serviceKey": KMA_API_Key, "pageNo": "1", "numOfRows": "100",
                "dataType": "JSON", "dataCd": "ASOS", "dateCd": "DAY",
                "startDt": Start_Date.strftime("%Y%m%d"),
                "endDt":   Today.strftime("%Y%m%d"), "stnIds": "131",
            }
            R = requests.get(
                "https://apis.data.go.kr/1360000/AsosDalyInfoService/getWthrDataList",
                params=Params, timeout=10
            )
            Items = R.json().get("response", {}).get("body", {}).get("items", {}).get("item", [])
            if Items:
                Data = []
                for Item in Items:
                    MT = float(Item.get("avgTa", 15) or 15)
                    Sol = float(Item.get("sumGsr", 12) or 12)
                    Data.append({
                        "날짜": Item.get("tm", ""), "출처": "기상청API",
                        "평균온도_C": MT,
                        "최고온도_C": float(Item.get("maxTa", MT+3) or MT+3),
                        "최저온도_C": float(Item.get("minTa", MT-3) or MT-3),
                        "상대습도_Pct": float(Item.get("avgRhm", 60) or 60),
                        "강수량_mm":   float(Item.get("sumRn", 0) or 0),
                        "풍속_ms":     float(Item.get("avgWs", 1.5) or 1.5),
                        "일사량_MJ":   Sol,
                        "증발산량ET0_mm": _calc_et0(MT, 6, Sol),
                        "데이터품질_Flag": "G",
                    })
                Source = "기상청API"
                Log_Lines.append(f"기상청 API 수집 성공: {len(Data)}건")
        except Exception as E:
            Log_Lines.append(f"기상청 API 실패: {E}")

    # ◆ WAMIS 크롤링 시도
    if Data is None and _REQUESTS_OK:
        try:
            Headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120"}
            R = requests.get(
                "https://www.wamis.go.kr/wkw/wk_flw_slct.aspx",
                params={"obscd": "131",
                        "startdt": Start_Date.strftime("%Y%m%d"),
                        "enddt":   Today.strftime("%Y%m%d"),
                        "gubun": "D"},
                headers=Headers, timeout=10
            )
            Soup  = BeautifulSoup(R.text, "html.parser")
            Table = Soup.find("table")
            if Table:
                Rows = Table.find_all("tr")
                Tmp  = []
                for Row in Rows[2:]:
                    Cells = Row.find_all("td")
                    if len(Cells) < 4:
                        continue
                    MT  = float(Cells[2].get_text(strip=True) or 15)
                    Sol = float(Cells[5].get_text(strip=True) or 12) if len(Cells) > 5 else 12.0
                    Tmp.append({
                        "날짜": Cells[0].get_text(strip=True), "출처": "WAMIS크롤링",
                        "평균온도_C": MT,
                        "최고온도_C": MT + 3,
                        "최저온도_C": MT - 3,
                        "상대습도_Pct": float(Cells[3].get_text(strip=True) or 60),
                        "강수량_mm":   float(Cells[1].get_text(strip=True) or 0),
                        "풍속_ms":     float(Cells[4].get_text(strip=True) or 1.5) if len(Cells) > 4 else 1.5,
                        "일사량_MJ":   Sol,
                        "증발산량ET0_mm": _calc_et0(MT, 6, Sol),
                        "데이터품질_Flag": "W",
                    })
                if Tmp:
                    Data   = Tmp
                    Source = "WAMIS크롤링"
                    Log_Lines.append(f"WAMIS 크롤링 성공: {len(Data)}건")
        except Exception as E:
            Log_Lines.append(f"WAMIS 크롤링 실패: {E}")

    # ◆ 시뮬레이션 Fallback
    if Data is None:
        Data = _sim_weather(Days, Season, Start_Date)
        Log_Lines.append(f"시뮬레이션 생성: {len(Data)}건 (Season={Season})")

    # CSV 저장
    Saved = ""
    try:
        os.makedirs(Save_Dir, exist_ok=True)
        Ts   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        Path = os.path.join(Save_Dir, f"weather_raw_{Ts}.csv")
        Keys = list(Data[0].keys()) if Data else []
        with open(Path, "w", newline="", encoding="utf-8-sig") as F:
            W = csv.DictWriter(F, fieldnames=Keys)
            W.writeheader(); W.writerows(Data)
        Saved = Path
        Log_Lines.append(f"Raw CSV 저장: {Path}")
    except Exception as E:
        Log_Lines.append(f"CSV 저장 오류: {E}")

    return {"data": Data, "source": Source, "saved_path": Saved, "log": " | ".join(Log_Lines)}


# ══════════════════════════════════════════════════════════
#  섹션 C — 토양 데이터 생성
# ══════════════════════════════════════════════════════════

def generate_soil(
    Weather_List: list[dict],
    Crop_Type: str,
    Initial_Moisture: float = 65.0,
    Save_Dir: str = "data/raw",
) -> dict:
    """
    [INPUT]  Weather_List, Crop_Type, Initial_Moisture, Save_Dir
    [OUTPUT] {"data": [...], "saved_path": str}

    물 수지 모델로 토양 수분 시뮬레이션.
    """
    Kc       = CROP_KC.get(Crop_Type, 1.0)
    Moisture = Initial_Moisture
    Result   = []
    for W in Weather_List:
        Rain_E   = W.get("강수량_mm", 0) * 0.6
        ET_E     = W.get("증발산량ET0_mm", 0) * Kc * 3.0
        Drain    = max(0.0, (Moisture - 90) * 0.3)
        Moisture = max(5.0, min(100.0, Moisture + Rain_E - ET_E - Drain + random.uniform(-1.5, 1.5)))
        EC_Base  = 0.8 + (Moisture / 100) * 1.5
        Result.append({
            "날짜":          W.get("날짜", ""),
            "토양수분_Pct":  round(Moisture, 1),
            "토양온도_C":    round(W.get("평균온도_C", 15) * 0.85 + random.uniform(-1, 1), 1),
            "전기전도도_dSm": round(min(4.0, max(0.1, EC_Base + random.uniform(-0.3, 0.3))), 2),
            "pH":            round(random.uniform(5.5, 7.5), 1),
        })

    Saved = ""
    try:
        os.makedirs(Save_Dir, exist_ok=True)
        Ts   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        Path = os.path.join(Save_Dir, f"soil_raw_{Ts}.csv")
        with open(Path, "w", newline="", encoding="utf-8-sig") as F:
            W2 = csv.DictWriter(F, fieldnames=list(Result[0].keys()))
            W2.writeheader(); W2.writerows(Result)
        Saved = Path
    except Exception:
        pass

    return {"data": Result, "saved_path": Saved}


# ══════════════════════════════════════════════════════════
#  섹션 D — 데이터 전처리 · 품질 관리
# ══════════════════════════════════════════════════════════

def _impute(Data: list[dict], Col: str, Idx: int) -> float:
    """결측값 선형 보간. 내부 전용."""
    Prev = Next = None
    for J in range(Idx - 1, -1, -1):
        V = Data[J].get(Col)
        if V not in (None, ""):
            try: Prev = float(V); break
            except (ValueError, TypeError): pass
    for J in range(Idx + 1, len(Data)):
        V = Data[J].get(Col)
        if V not in (None, ""):
            try: Next = float(V); break
            except (ValueError, TypeError): pass
    if Prev is not None and Next is not None:
        return round((Prev + Next) / 2, 2)
    if Prev is not None: return Prev
    if Next is not None: return Next
    Vals = [float(R[Col]) for R in Data if R.get(Col) not in (None, "")]
    return round(statistics.mean(Vals), 2) if Vals else 0.0


def preprocess(
    Raw_Data: list[dict],
    Save_Dir: str = "data/processed",
) -> dict:
    """
    [INPUT]  Raw_Data (기상 또는 토양 리스트), Save_Dir
    [OUTPUT] {
        "data":           전처리된 리스트,
        "quality_score":  float (0~100),
        "quality_grade":  str (A/B/C/D),
        "quality_desc":   str,
        "total_missing":  int,
        "total_outliers": int,
        "source_counts":  dict,
        "saved_csv":      str,
        "saved_xlsx":     str,
        "log":            str,
    }

    1. 결측값 탐지 → 선형 보간
    2. 이상값 탐지 → 범위 클리핑
    3. 논리 검증 (최고 >= 평균 >= 최저)
    4. 품질 점수 A/B/C/D
    5. CSV (LLM용) + Excel (수치조정용) 저장
    """
    Num_Cols      = list(VALID_RANGE.keys())
    Processed     = []
    Total_Missing = 0
    Total_Out     = 0
    Source_Counts: dict = {}

    for Idx, Row in enumerate(Raw_Data):
        Source_Counts[Row.get("데이터품질_Flag", "S")] = \
            Source_Counts.get(Row.get("데이터품질_Flag", "S"), 0) + 1
        C = dict(Row)
        Miss_Cnt = Out_Cnt = 0

        # 결측값 처리
        for Col in Num_Cols:
            if C.get(Col) in (None, ""):
                C[Col] = _impute(Raw_Data, Col, Idx); Miss_Cnt += 1

        # 이상값 클리핑
        for Col, (Lo, Hi) in VALID_RANGE.items():
            try:
                V = float(C.get(Col, 0) or 0)
                if V < Lo or V > Hi:
                    C[Col] = round(max(Lo, min(Hi, V)), 2); Out_Cnt += 1
            except (ValueError, TypeError):
                C[Col] = Lo; Out_Cnt += 1

        # 기온 논리 검증
        try:
            MX = float(C.get("최고온도_C", 0) or 0)
            MT = float(C.get("평균온도_C", 0) or 0)
            MN = float(C.get("최저온도_C", 0) or 0)
            if not (MX >= MT >= MN):
                C["최고온도_C"] = round(MT + abs(MX - MT), 1)
                C["최저온도_C"] = round(MT - abs(MT - MN), 1)
                Out_Cnt += 1
        except (ValueError, TypeError):
            pass

        C["전처리_Flag"]      = "P" if (Miss_Cnt + Out_Cnt) > 0 else "OK"
        C["결측처리_Count"]   = Miss_Cnt
        C["이상값처리_Count"] = Out_Cnt
        Processed.append(C)
        Total_Missing += Miss_Cnt; Total_Out += Out_Cnt

    # 품질 점수 산출
    N = len(Processed) or 1
    Score = 100.0
    Score -= (Total_Missing / (N * len(Num_Cols))) * 30
    Score -= (Total_Out     / (N * len(Num_Cols))) * 20
    Score -= (Source_Counts.get("S", 0) / N) * 20
    Score  = round(max(0.0, min(100.0, Score)), 1)
    Grade  = "D"
    for G, T in [("A", 90), ("B", 70), ("C", 50), ("D", 0)]:
        if Score >= T: Grade = G; break
    Desc   = {"A": "우수 — 바로 사용 가능", "B": "양호 — 일부 주의",
               "C": "보통 — 추가 수집 권장", "D": "불량 — 재수집 필요"}.get(Grade, "")

    # 저장
    Saved_CSV = Saved_XLSX = ""
    try:
        os.makedirs(Save_Dir, exist_ok=True)
        Ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        # CSV — LLM 분석용
        Saved_CSV = os.path.join(Save_Dir, f"processed_{Ts}.csv")
        Keys = list(Processed[0].keys()) if Processed else []
        with open(Saved_CSV, "w", newline="", encoding="utf-8-sig") as F:
            W = csv.DictWriter(F, fieldnames=Keys, extrasaction="ignore")
            W.writeheader(); W.writerows(Processed)
        # Excel — 수치 조정용
        if _EXCEL_OK:
            Saved_XLSX = os.path.join(Save_Dir, f"processed_{Ts}.xlsx")
            WB = openpyxl.Workbook()
            WS = WB.active
            WS.title = "전처리데이터"
            WS.append(Keys)
            for R in Processed:
                WS.append([R.get(K, "") for K in Keys])
            WB.save(Saved_XLSX)
    except Exception as E:
        Saved_CSV = Saved_XLSX = f"저장오류: {E}"

    return {
        "data":           Processed,
        "quality_score":  Score,
        "quality_grade":  Grade,
        "quality_desc":   Desc,
        "total_missing":  Total_Missing,
        "total_outliers": Total_Out,
        "source_counts":  Source_Counts,
        "saved_csv":      Saved_CSV,
        "saved_xlsx":     Saved_XLSX,
        "log":            f"전처리 완료 {N}행 | 품질 {Grade} ({Score}점)",
    }


# ══════════════════════════════════════════════════════════
#  섹션 E — 모델 개발 · 검증
# ══════════════════════════════════════════════════════════

class _RuleWeightedModel:
    """가중치 규칙 기반 관수 예측 모델. 내부 전용."""
    def __init__(self, Crop_Type: str):
        self.Crop_Type   = Crop_Type
        self.Weights     = [0.40, 0.15, 0.10, 0.15, 0.10, 0.05, 0.03, 0.02]
        self.Train_Hist  = []

    def _predict_one(self, FV: list[float], NMin: dict, NMax: dict) -> int:
        T = CROP_THRESHOLDS.get(self.Crop_Type, CROP_THRESHOLDS["상추/채소"])
        S_Range = NMax.get("토양수분_Pct", 100) - NMin.get("토양수분_Pct", 0)
        Soil    = FV[0] * S_Range + NMin.get("토양수분_Pct", 0)
        R_Range = NMax.get("강수량_mm", 100)   - NMin.get("강수량_mm", 0)
        Rain    = FV[3] * R_Range
        if Soil < T["Min"]:               return 0
        elif Soil < T["Opt"] * 0.85:      return 2 if Rain > 20 else 1
        elif Soil > T["Max"]:             return 3
        else:                              return 2

    def train(self, FM: list, Labels: list, NMin: dict, NMax: dict, Epochs: int = 10):
        LR = 0.01
        for _ in range(Epochs):
            Correct = 0
            for FV, TL in zip(FM, Labels):
                PL = self._predict_one(FV, NMin, NMax)
                if PL == TL: Correct += 1
                else:
                    Err = float(TL - PL) * LR
                    self.Weights = [W + Err * V for W, V in zip(self.Weights, FV)]
                    TW = sum(abs(W) for W in self.Weights)
                    if TW > 0: self.Weights = [W / TW for W in self.Weights]
            self.Train_Hist.append(round(Correct / max(len(Labels), 1) * 100, 1))

    def predict_batch(self, FM: list, NMin: dict, NMax: dict) -> list[int]:
        return [self._predict_one(FV, NMin, NMax) for FV in FM]


def _norm(Data: list[dict]) -> tuple[list, dict, dict]:
    """Min-Max 정규화. 내부 전용."""
    NMin = {}; NMax = {}
    for Col in FEATURE_COLS:
        Vals = [float(R[Col]) for R in Data if R.get(Col) not in (None, "")]
        NMin[Col] = min(Vals) if Vals else 0.0
        NMax[Col] = max(Vals) if Vals else 1.0
        if NMax[Col] == NMin[Col]: NMax[Col] = NMin[Col] + 1.0
    FM = []
    for R in Data:
        Vec = []
        for Col in FEATURE_COLS:
            try: V = float(R.get(Col) or NMin[Col])
            except (ValueError, TypeError): V = NMin[Col]
            Vec.append(round((V - NMin[Col]) / (NMax[Col] - NMin[Col]), 4))
        FM.append(Vec)
    return FM, NMin, NMax


def _labels(Data: list[dict], Crop_Type: str) -> list[int]:
    """레이블 생성. 내부 전용."""
    T = CROP_THRESHOLDS.get(Crop_Type, CROP_THRESHOLDS["상추/채소"])
    Out = []
    for R in Data:
        M = float(R.get("토양수분_Pct", 65) or 65)
        Rain = float(R.get("최근3일강수_mm", 0) or 0)
        if M < T["Min"]:            Out.append(0)
        elif M < T["Opt"] * 0.85:  Out.append(2 if Rain > 20 else 1)
        elif M > T["Max"]:          Out.append(3)
        else:                        Out.append(2)
    return Out


def _augment(Data: list[dict]) -> list[dict]:
    """파생 피처 추가. 내부 전용."""
    Aug = []
    for I, R in enumerate(Data):
        NR = dict(R)
        NR["최근3일강수_mm"]   = round(sum(float(Data[J].get("강수량_mm", 0) or 0) for J in range(max(0, I-2), I+1)), 1)
        Window = [float(Data[J].get("토양수분_Pct", 65) or 65) for J in range(max(0, I-2), I+1)]
        NR["이동평균수분_Pct"] = round(statistics.mean(Window), 1)
        Aug.append(NR)
    return Aug


def _calc_metrics(True_L: list[int], Pred_L: list[int]) -> dict:
    """분류 성능 지표 계산. 내부 전용."""
    N = 4
    CM = [[0]*N for _ in range(N)]
    for T, P in zip(True_L, Pred_L):
        if 0 <= T < N and 0 <= P < N: CM[T][P] += 1
    Acc = sum(CM[I][I] for I in range(N)) / max(len(True_L), 1)
    Class_M = {}
    for I in range(N):
        TP = CM[I][I]
        FP = sum(CM[J][I] for J in range(N)) - TP
        FN = sum(CM[I][J] for J in range(N)) - TP
        Prec = TP / max(TP + FP, 1)
        Rec  = TP / max(TP + FN, 1)
        F1   = 2 * Prec * Rec / max(Prec + Rec, 1e-9)
        Class_M[IRRIGATION_LABELS[I]] = {
            "Precision": round(Prec, 3), "Recall": round(Rec, 3),
            "F1": round(F1, 3), "Support": sum(CM[I]),
        }
    Macro_F1 = statistics.mean(V["F1"] for V in Class_M.values())
    return {"accuracy": round(Acc, 3), "macro_f1": round(Macro_F1, 3),
            "class_metrics": Class_M, "confusion_matrix": CM}


def train_and_validate(
    Weather_Data: list[dict],
    Soil_Data: list[dict],
    Crop_Type: str,
    Save_Dir: str = "data/model",
) -> dict:
    """
    [INPUT]  Weather_Data, Soil_Data, Crop_Type, Save_Dir
    [OUTPUT] {
        "model":        _RuleWeightedModel 객체 (Frontend로 전달),
        "norm_min":     dict,
        "norm_max":     dict,
        "cv_result":    dict (교차검증 결과),
        "final_metrics":dict (최종 성능),
        "train_history":list[float],
        "label_dist":   dict,
        "log":          str,
    }

    파이프라인:
    ① 병합 → ② 파생 피처 → ③ 레이블 → ④ 정규화
    → ⑤ Train/Test 분할 → ⑥ 학습 → ⑦ K-Fold 검증
    → ⑧ 최종 평가 → ⑨ 모델 저장
    """
    # ① 병합
    Merged = []
    Soil_Map = {S["날짜"]: S for S in Soil_Data}
    for W in Weather_Data:
        S = Soil_Map.get(W.get("날짜", ""), {})
        if not S and len(Soil_Data) == len(Weather_Data):
            S = Soil_Data[len(Merged)]
        Merged.append({**W, **S})

    # ② 파생 피처
    Aug    = _augment(Merged)
    Labels = _labels(Aug, Crop_Type)
    Dist   = {IRRIGATION_LABELS[K]: V for K, V in Counter(Labels).items()}

    # ④ 정규화
    FM, NMin, NMax = _norm(Aug)

    # ⑤ Train/Test 분할 (80:20)
    Split = max(1, int(len(Aug) * 0.8))
    TrF, TrL = FM[:Split], Labels[:Split]
    TeF, TeL = FM[Split:], Labels[Split:]

    # ⑥ 학습
    Model = _RuleWeightedModel(Crop_Type)
    if TrF: Model.train(TrF, TrL, NMin, NMax)

    # ⑦ K-Fold 교차 검증 (K=3)
    K    = 3
    N    = len(Aug)
    Idxs = list(range(N)); random.shuffle(Idxs)
    Fold_Accs = []; Fold_F1s = []
    for Ki in range(K):
        VStart = Ki * (N // K)
        VEnd   = VStart + (N // K) if Ki < K - 1 else N
        Val_Idx  = Idxs[VStart:VEnd]
        Tr_Idx   = [I for I in Idxs if I not in set(Val_Idx)]
        FM_fold  = _RuleWeightedModel(Crop_Type)
        if Tr_Idx: FM_fold.train([FM[I] for I in Tr_Idx], [Labels[I] for I in Tr_Idx], NMin, NMax, Epochs=3)
        Preds  = FM_fold.predict_batch([FM[I] for I in Val_Idx], NMin, NMax)
        M      = _calc_metrics([Labels[I] for I in Val_Idx], Preds)
        Fold_Accs.append(M["accuracy"]); Fold_F1s.append(M["macro_f1"])

    CV = {
        "mean_accuracy": round(statistics.mean(Fold_Accs), 3),
        "mean_f1":       round(statistics.mean(Fold_F1s), 3),
        "std_accuracy":  round(statistics.stdev(Fold_Accs) if len(Fold_Accs) > 1 else 0.0, 3),
        "method":        f"{K}-Fold 교차검증",
    }

    # ⑧ 최종 평가
    Final_P = Model.predict_batch(TeF if TeF else FM, NMin, NMax)
    Final_M = _calc_metrics(TeL if TeL else Labels, Final_P)

    # ⑨ 저장
    try:
        os.makedirs(Save_Dir, exist_ok=True)
        Ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        with open(os.path.join(Save_Dir, f"weights_{Ts}.csv"), "w", newline="", encoding="utf-8-sig") as F:
            W = csv.writer(F)
            W.writerow(["피처", "가중치"])
            for Col, Wt in zip(FEATURE_COLS, Model.Weights): W.writerow([Col, round(Wt, 6)])
        with open(os.path.join(Save_Dir, f"cv_report_{Ts}.json"), "w", encoding="utf-8") as F:
            json.dump({"cv": CV, "final": Final_M, "crop": Crop_Type, "ts": Ts}, F, ensure_ascii=False, indent=2)
    except Exception as E:
        pass

    return {
        "model":         Model,
        "norm_min":      NMin,
        "norm_max":      NMax,
        "cv_result":     CV,
        "final_metrics": Final_M,
        "train_history": Model.Train_Hist,
        "label_dist":    Dist,
        "log":           f"모델 학습 완료 | CV F1={CV['mean_f1']:.3f} | 정확도={Final_M['accuracy']:.1%}",
    }


# ══════════════════════════════════════════════════════════
#  섹션 F — 관수 의사결정
# ══════════════════════════════════════════════════════════

def make_decision(
    Model:         "_RuleWeightedModel",
    Norm_Min:      dict,
    Norm_Max:      dict,
    Weather_Today: dict,
    Soil_Today:    dict,
    Weather_Hist:  list[dict],
    Crop_Type:     str,
    Area_m2:       float,
) -> dict:
    """
    [INPUT]  모델·정규화값·오늘 기상·토양·이력·작물·면적
    [OUTPUT] {
        "status":          str  (즉시관수/관수권장/적정상태/과습주의),
        "urgency":         int  (0~5),
        "reason":          str,
        "soil_moisture":   float,
        "threshold_min":   float,
        "threshold_opt":   float,
        "threshold_max":   float,
        "etc_mm":          float,
        "deficit_mm":      float,
        "volume_l":        float,
        "recent_rain_mm":  float,
        "heat_stress":     bool,
        "area_m2":         float,
        "label_int":       int,
    }
    """
    T    = CROP_THRESHOLDS.get(Crop_Type, CROP_THRESHOLDS["상추/채소"])
    Kc   = CROP_KC.get(Crop_Type, 1.0)
    Rain3 = sum(W.get("강수량_mm", 0) for W in Weather_Hist[-3:])
    Mois  = Soil_Today.get("토양수분_Pct", 65)

    Single = {**Weather_Today, **Soil_Today,
              "최근3일강수_mm": Rain3,
              "이동평균수분_Pct": Mois}
    FM, _, _ = _norm([Single])
    Label = Model._predict_one(FM[0], Norm_Min, Norm_Max) if FM else 2
    Status = IRRIGATION_LABELS.get(Label, "적정상태")

    ET0   = Weather_Today.get("증발산량ET0_mm", 0)
    ETC   = ET0 * Kc
    Eff_R = min(Weather_Today.get("강수량_mm", 0) * 0.8, ETC)
    Def   = max(0.0, ETC - Eff_R)
    Vol   = round(Def * Area_m2, 1)

    Reasons = {
        "즉시관수": f"토양 수분 {Mois}%가 최소 기준({T['Min']}%) 미달",
        "관수권장": f"토양 수분 {Mois}%가 최적({T['Opt']}%) 이하",
        "적정상태": f"토양 수분 {Mois}%가 적정 범위 유지 중",
        "과습주의": f"토양 수분 {Mois}%가 최대({T['Max']}%) 초과",
    }
    Urgency = {0: 5, 1: 3, 2: 0, 3: 2}

    return {
        "status":         Status,
        "urgency":        Urgency.get(Label, 0),
        "reason":         Reasons.get(Status, ""),
        "soil_moisture":  Mois,
        "threshold_min":  T["Min"],
        "threshold_opt":  T["Opt"],
        "threshold_max":  T["Max"],
        "etc_mm":         round(ETC, 2),
        "deficit_mm":     round(Def, 2),
        "volume_l":       Vol,
        "recent_rain_mm": round(Rain3, 1),
        "heat_stress":    Weather_Today.get("평균온도_C", 0) > 33,
        "area_m2":        Area_m2,
        "label_int":      Label,
    }


# ══════════════════════════════════════════════════════════
#  섹션 G — AI 권고문 생성
# ══════════════════════════════════════════════════════════

def get_advice(
    Decision:      dict,
    Weather_Today: dict,
    Soil_Today:    dict,
    Crop_Type:     str,
    API_Key:       str | None = None,
) -> dict:
    """
    [INPUT]  의사결정 결과, 기상·토양 데이터, 작물, API 키
    [OUTPUT] {"advice": str, "source": "claude_ai" | "rule_based"}
    """
    import urllib.request, urllib.error

    Data = {
        "crop":         Crop_Type,
        "moisture":     Decision["soil_moisture"],
        "opt_moisture": Decision["threshold_opt"],
        "status":       Decision["status"],
        "temp":         Weather_Today.get("평균온도_C", 20),
        "humid":        Weather_Today.get("상대습도_Pct", 60),
        "rain":         Weather_Today.get("강수량_mm", 0),
        "rain_3day":    Decision["recent_rain_mm"],
        "et0":          Weather_Today.get("증발산량ET0_mm", 0),
        "deficit":      Decision["deficit_mm"],
        "volume":       Decision["volume_l"],
        "heat":         Decision["heat_stress"],
        "ph":           Soil_Today.get("pH", 6.5),
        "ec":           Soil_Today.get("전기전도도_dSm", 1.0),
    }

    if API_Key and _REQUESTS_OK:
        try:
            import json as _json
            Payload = _json.dumps({
                "model": "claude-sonnet-4-20250514", "max_tokens": 600,
                "system": "당신은 스마트 농업 전문 컨설턴트입니다. 주어진 기상·토양 데이터를 분석하여 농업인에게 실용적이고 구체적인 관수 권고문을 3~5문장으로 작성하세요. 수치를 반드시 포함하세요.",
                "messages": [{"role": "user", "content":
                    f"작물:{Data['crop']}, 토양수분:{Data['moisture']}%(적정:{Data['opt_moisture']}%), "
                    f"판단:{Data['status']}, 기온:{Data['temp']}°C, 습도:{Data['humid']}%, "
                    f"강수:{Data['rain']}mm(3일:{Data['rain_3day']}mm), ET0:{Data['et0']}mm, "
                    f"부족량:{Data['deficit']}mm, 권장관수:{Data['volume']}L, "
                    f"고온스트레스:{'있음' if Data['heat'] else '없음'}, "
                    f"pH:{Data['ph']}, EC:{Data['ec']}dS/m\n\n위 데이터로 관수 권고문 작성해주세요."
                }]
            }).encode("utf-8")
            Req = urllib.request.Request(
                "https://api.anthropic.com/v1/messages", data=Payload,
                headers={"Content-Type": "application/json",
                         "x-api-key": API_Key, "anthropic-version": "2023-06-01"},
                method="POST"
            )
            with urllib.request.urlopen(Req, timeout=15) as Resp:
                import json as _j2
                Txt = _j2.loads(Resp.read())["content"][0]["text"]
                return {"advice": Txt, "source": "claude_ai"}
        except Exception:
            pass

    # 규칙 기반 Fallback
    Lines = []
    S = Data["status"]
    if S == "즉시관수":
        Lines.append(f"⚠ {Data['crop']} 재배지의 토양 수분 {Data['moisture']}%가 심각하게 부족합니다. 즉시 약 {Data['volume']:.0f}L의 관수가 필요합니다.")
        Lines.append("이른 아침(06~08시) 또는 저녁(18~20시)에 점적관수 방식으로 공급하면 증발 손실을 최소화할 수 있습니다.")
        if Data["heat"]: Lines.append(f"기온 {Data['temp']}°C로 고온 스트레스 구간 — 잎 직접 살수는 화상 위험이 있으므로 지표 관수를 권장합니다.")
    elif S == "관수권장":
        Lines.append(f"{Data['crop']}의 토양 수분 {Data['moisture']}%는 적정보다 다소 낮습니다. 오늘 {Data['volume']:.0f}L 정도 보충을 권장합니다.")
        if Data["rain_3day"] > 10: Lines.append(f"최근 3일 강수 {data['rain_3day']}mm 감안하여 권장량의 60~70%만 관수해도 충분할 수 있습니다.")
    elif S == "적정상태":
        Lines.append(f"{Data['crop']} 재배지 토양 수분 {Data['moisture']}%는 현재 최적 범위입니다. 오늘 관수는 불필요합니다.")
        Lines.append("이틀 후 토양 수분을 재점검하고 기상 예보를 확인하세요.")
    else:  # 과습주의
        Lines.append(f"토양 수분 {Data['moisture']}%로 과습 상태입니다. 관수를 즉시 중단하고 배수로를 점검해 주세요.")
        Lines.append("과습이 지속되면 뿌리 호흡이 방해되어 생육 불량·병해 발생 위험이 높아집니다.")
    if Data["ph"] < 5.8: Lines.append(f"토양 pH {Data['ph']}로 산성 — 관수 시 석회 처리를 병행하세요.")
    if Data["ec"] > 2.0: Lines.append(f"EC {Data['ec']} dS/m로 염류 집적 가능성 — 충분한 관수로 용탈하세요.")

    return {"advice": "\n".join(Lines), "source": "rule_based"}


# ══════════════════════════════════════════════════════════
#  섹션 H — 이력 관리
# ══════════════════════════════════════════════════════════

def load_history(File_Path: str = "irrigation_history.json") -> list[dict]:
    """
    [INPUT]  File_Path
    [OUTPUT] 관수 이력 리스트
    """
    if not os.path.exists(File_Path): return []
    try:
        with open(File_Path, "r", encoding="utf-8") as F:
            Data = json.load(F)
            return Data if isinstance(Data, list) else []
    except Exception: return []


def save_history(Records: list[dict], File_Path: str = "irrigation_history.json") -> bool:
    """
    [INPUT]  Records 리스트, File_Path
    [OUTPUT] 성공 여부 (bool)
    """
    try:
        with open(File_Path, "w", encoding="utf-8") as F:
            json.dump(Records[-60:], F, ensure_ascii=False, indent=2)
        return True
    except Exception: return False


def add_history_record(
    Records:        list[dict],
    Date_Str:       str,
    Crop_Type:      str,
    Decision:       dict,
    Irrigated:      bool,
    Actual_Vol_L:   float = 0.0,
    File_Path:      str   = "irrigation_history.json",
) -> list[dict]:
    """
    [INPUT]  기존 이력 + 신규 레코드 데이터
    [OUTPUT] 업데이트된 이력 리스트
    """
    Records.append({
        "날짜": Date_Str, "작물": Crop_Type,
        "토양수분_Pct": Decision["soil_moisture"],
        "판단결과": Decision["status"],
        "긴급도": Decision["urgency"],
        "관수여부": Irrigated,
        "실제관수량_L": Actual_Vol_L,
        "권장관수량_L": Decision["volume_l"],
    })
    save_history(Records, File_Path)
    return Records


def calc_stats(Records: list[dict]) -> dict:
    """
    [INPUT]  이력 리스트
    [OUTPUT] {
        "total": int, "irrigated": int,
        "total_vol_L": float, "saved_vol_L": float,
        "status_dist": dict,
    }
    """
    Irr   = [R for R in Records if R.get("관수여부")]
    Dist  = {}
    for R in Records:
        S = R.get("판단결과", "알 수 없음")
        Dist[S] = Dist.get(S, 0) + 1
    Total_Vol = sum(R.get("실제관수량_L", 0) for R in Irr)
    Rec_Vol   = sum(R.get("권장관수량_L", 0) for R in Records)
    return {
        "total": len(Records), "irrigated": len(Irr),
        "total_vol_L": Total_Vol,
        "saved_vol_L": max(0, Rec_Vol - Total_Vol),
        "status_dist": Dist,
    }


# ══════════════════════════════════════════════════════════
#  섹션 I — 데이터 내보내기
# ══════════════════════════════════════════════════════════

def export_data(
    Weather_List:    list[dict],
    Soil_List:       list[dict],
    History_Records: list[dict],
    Base_Name:       str = "export",
    Save_Dir:        str = "data/processed",
) -> dict:
    """
    [INPUT]  기상·토양·이력 리스트, 파일 이름 기반, 저장 폴더
    [OUTPUT] {"csv_paths": [...], "xlsx_paths": [...], "log": str}

    CSV  — LLM 분석에 최적 (구분자 기반 텍스트)
    Excel — 수치 조정 · 시각화에 최적 (셀 서식 포함)
    """
    os.makedirs(Save_Dir, exist_ok=True)
    Ts        = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    CSV_Paths = []; XLSX_Paths = []; Logs = []

    Datasets = [
        (Weather_List,    f"{Base_Name}_weather_{Ts}"),
        (Soil_List,       f"{Base_Name}_soil_{Ts}"),
        (History_Records, f"{Base_Name}_history_{Ts}"),
    ]
    for Data, Name in Datasets:
        if not Data: continue
        Keys = list(Data[0].keys())
        # CSV
        try:
            CP = os.path.join(Save_Dir, f"{Name}.csv")
            with open(CP, "w", newline="", encoding="utf-8-sig") as F:
                W = csv.DictWriter(F, fieldnames=Keys, extrasaction="ignore")
                W.writeheader(); W.writerows(Data)
            CSV_Paths.append(CP)
        except Exception as E:
            Logs.append(f"CSV 오류 [{Name}]: {E}")
        # Excel
        if _EXCEL_OK:
            try:
                XP = os.path.join(Save_Dir, f"{Name}.xlsx")
                WB = openpyxl.Workbook(); WS = WB.active; WS.title = "데이터"
                WS.append(Keys)
                for R in Data: WS.append([R.get(K, "") for K in Keys])
                WB.save(XP); XLSX_Paths.append(XP)
            except Exception as E:
                Logs.append(f"Excel 오류 [{Name}]: {E}")

    return {
        "csv_paths":  CSV_Paths,
        "xlsx_paths": XLSX_Paths,
        "log":        f"CSV {len(CSV_Paths)}개, Excel {len(XLSX_Paths)}개 저장 완료" + ((" | 오류: " + ", ".join(Logs)) if Logs else ""),
    }
