"""
==============================================================
  ui_frontend.py  —  FRONTEND (UI Layer)
==============================================================

  [역할]
  "화면으로 보이는 코딩" — 사용자와의 모든 상호작용 담당.
  데이터 연산·저장이 단 한 줄도 없습니다.
  Backend(backend_logic.py)의 함수를 호출하고 결과를 표시합니다.

  [Input / Output 규약]
  Input  : 키보드 입력 (Safe_Input, Safe_Float)
  Output : 터미널 출력 (대시보드, 차트, 표, 메시지)

  [Frontend가 Backend를 부르는 방식]
  from backend_logic import collect_weather, preprocess, ...
  result = collect_weather(Days=30, Season="봄")   # ← Backend 호출
  Print_Weather_Panel(result["data"][-1])           # ← Frontend 출력

  실행: python ui_frontend.py
==============================================================
"""

import os
import sys
import datetime

# ── Backend 전체를 단일 import ────────────────────────────
try:
    from backend_logic import (
        CROP_THRESHOLDS, SEASON_PARAMS,
        collect_weather, generate_soil,
        preprocess,
        train_and_validate,
        make_decision,
        get_advice,
        load_history, save_history, add_history_record, calc_stats,
        export_data,
        IRRIGATION_LABELS,
    )
except ImportError as E:
    print(f"\n[오류] backend_logic.py 를 찾을 수 없습니다: {E}")
    print("  → ui_frontend.py 와 backend_logic.py 가 같은 폴더에 있어야 합니다.")
    sys.exit(1)


# ══════════════════════════════════════════════════════════
#  섹션 A — 색상 상수 (UI 전용)
# ══════════════════════════════════════════════════════════

RST  = "\033[0m";  BOLD = "\033[1m"
CYAN = "\033[96m"; GRN  = "\033[92m"; YLW = "\033[93m"
RED  = "\033[91m"; BLU  = "\033[94m"; GRY = "\033[90m"
WHT  = "\033[97m"

STATUS_CFG = {
    "즉시관수": (RED,  "🚨"),
    "관수권장": (YLW,  "⚠️ "),
    "적정상태": (GRN,  "✅"),
    "과습주의": (BLU,  "💧"),
}


# ══════════════════════════════════════════════════════════
#  섹션 B — Input 헬퍼 (사용자 입력)
# ══════════════════════════════════════════════════════════

def Safe_Input(Prompt: str, Default: str = "") -> str:
    """
    [Input]  키보드 입력
    [Output] 입력 문자열 (빈 입력 시 Default 반환)
    KeyboardInterrupt 처리 포함.
    """
    try:
        Disp  = f"  {Prompt}" + (f" [{Default}]" if Default else "") + ": "
        Val   = input(Disp).strip()
        return Val if Val else Default
    except KeyboardInterrupt:
        print(f"\n  {YLW}입력 취소{RST}")
        return Default


def Safe_Float(Prompt: str, Default: float) -> float:
    """
    [Input]  키보드 입력 (실수)
    [Output] float (잘못된 입력 시 Default)
    """
    Raw = Safe_Input(Prompt, str(Default))
    try:
        return float(Raw)
    except ValueError:
        Print_Msg(f"숫자 형식이 아닙니다 → 기본값 {Default} 사용", "warn")
        return Default


def Select_From_List(Title: str, Options: list[str], Default_Idx: int = 0) -> str:
    """
    [Input]  번호 선택
    [Output] 선택된 항목 문자열
    """
    Print_Section(Title)
    for I, Opt in enumerate(Options, 1):
        Mark = f"  {CYAN}← 현재 기본{RST}" if I - 1 == Default_Idx else ""
        print(f"  {I}. {Opt}{Mark}")
    while True:
        try:
            C = int(Safe_Input("번호 선택", str(Default_Idx + 1)))
            if 1 <= C <= len(Options):
                print(f"  {GRN}✔ 선택: {BOLD}{Options[C-1]}{RST}")
                return Options[C - 1]
        except ValueError:
            pass
        Print_Msg(f"1~{len(Options)} 사이의 번호를 입력하세요", "error")


# ══════════════════════════════════════════════════════════
#  섹션 C — Output 헬퍼 (공통 출력 컴포넌트)
# ══════════════════════════════════════════════════════════

def Print_Section(Title: str):
    """섹션 제목 출력"""
    print(f"\n  {BOLD}[ {Title} ]{RST}")


def Print_Msg(Text: str, Kind: str = "info"):
    """
    단일 메시지 출력.
    Kind: "info" | "success" | "warn" | "error" | "hint"
    """
    Color = {"info": CYAN, "success": GRN, "warn": YLW, "error": RED, "hint": GRY}.get(Kind, RST)
    Icon  = {"info": "ℹ", "success": "✅", "warn": "⚠️ ", "error": "❌", "hint": "·"}.get(Kind, "")
    print(f"  {Color}{Icon} {Text}{RST}")


def Print_Progress(Step: int, Total: int, Label: str):
    """진행 상태 바 출력"""
    Filled = int(Step / Total * 20)
    Bar    = f"{GRN}{'█' * Filled}{'░' * (20 - Filled)}{RST}"
    print(f"  {Bar} {Step}/{Total} {Label}")


# ══════════════════════════════════════════════════════════
#  섹션 D — Output 컴포넌트 (개별 패널)
# ══════════════════════════════════════════════════════════

def Print_Status_Panel(Decision: dict):
    """
    [Output] 관수 판단 결과 패널
    """
    Status = Decision["status"]
    Color, Symbol = STATUS_CFG.get(Status, (GRN, "✅"))
    Urgency_Stars  = "★" * Decision["urgency"] + "☆" * (5 - Decision["urgency"])

    print(f"\n  {BOLD}┌─ 관수 판단 결과 {'─'*42}┐{RST}")
    print(f"  │  {Color}{BOLD}{Symbol}  [ {Status} ]  {Symbol}{RST}")
    print(f"  │  {GRY}→ {Decision['reason']}{RST}")
    print(f"  │  {GRY}  긴급도: {Urgency_Stars}{RST}")
    if Decision["heat_stress"]:
        print(f"  │  {RED}  ⚠ 고온 스트레스 경고{RST}")
    print(f"  {BOLD}└{'─'*63}┘{RST}")


def Print_Moisture_Gauge(Decision: dict, Soil_Today: dict):
    """
    [Output] 토양 수분 게이지 + 부가 토양 정보 패널
    """
    Moisture  = Decision["soil_moisture"]
    Mn, Op, Mx = Decision["threshold_min"], Decision["threshold_opt"], Decision["threshold_max"]
    Width     = 42
    Gauge     = []
    for I in range(Width):
        Pct = I / Width * 100
        if Pct < Mn:   Gauge.append(f"{RED}▓{RST}")
        elif Pct < Op: Gauge.append(f"{YLW}▓{RST}")
        elif Pct <= Mx: Gauge.append(f"{GRN}▓{RST}")
        else:           Gauge.append(f"{BLU}▓{RST}")
    Arrow_Pos = min(int(Moisture / 100 * Width), Width - 1)
    Gauge[Arrow_Pos] = f"{WHT}{BOLD}▼{RST}"

    pH  = Soil_Today.get("pH", 6.5)
    EC  = Soil_Today.get("전기전도도_dSm", 1.0)
    ST  = Soil_Today.get("토양온도_C", 15)
    pH_C  = RED if pH < 5.5 or pH > 7.5 else YLW if pH < 6.0 or pH > 7.0 else GRN
    EC_C  = RED if EC > 2.5 else YLW if EC > 1.8 else GRN

    print(f"\n  {BOLD}┌─ 토양 수분 상태 {'─'*41}┐{RST}")
    print(f"  │  0%                              50%                       100%  │")
    print(f"  │  {''.join(Gauge)}  {Moisture}%  │")
    print(f"  │  {RED}▲{Mn}%{RST}          {GRN}▲{Op}%{RST}          {BLU}▲{Mx}%{RST}                    │")
    print(f"  │                                                               │")
    print(f"  │  🧪 pH: {pH_C}{BOLD}{pH:>4.1f}{RST}   ⚡ EC: {EC_C}{BOLD}{EC:>5.2f}dS/m{RST}   🌡 토양온도: {BOLD}{ST:>4.1f}°C{RST}  │")
    print(f"  {BOLD}└{'─'*63}┘{RST}")


def Print_Weather_Panel(Weather: dict):
    """
    [Output] 오늘의 기상 데이터 패널
    """
    T  = Weather.get("평균온도_C", 0)
    H  = Weather.get("상대습도_Pct", 0)
    R  = Weather.get("강수량_mm", 0)
    W  = Weather.get("풍속_ms", 0)
    S  = Weather.get("일사량_MJ", 0)
    E  = Weather.get("증발산량ET0_mm", 0)
    Source = Weather.get("출처", "-")

    TC = RED if T > 33 else YLW if T > 28 else GRN
    HC = BLU if H > 80 else GRN if H > 50 else YLW
    RC = CYAN if R > 0 else GRY

    print(f"\n  {BOLD}┌─ 오늘의 기상 데이터 {'─'*36}┐{RST}")
    print(f"  │  🌡 평균 온도: {TC}{BOLD}{T:>5.1f}°C{RST}      💧 습도: {HC}{BOLD}{H:>5.1f}%{RST}                    │")
    print(f"  │  🌧 강수량:   {RC}{BOLD}{R:>5.1f}mm{RST}      💨 풍속: {BOLD}{W:>5.1f}m/s{RST}                    │")
    print(f"  │  ☀ 일사량:   {BOLD}{S:>5.1f}MJ/m²{RST}  🌿 ET₀: {BOLD}{E:>5.2f}mm{RST}                       │")
    print(f"  │  {GRY}데이터 출처: {Source}{RST}")
    print(f"  {BOLD}└{'─'*63}┘{RST}")


def Print_Irrigation_Panel(Decision: dict):
    """
    [Output] 관수 권고량 패널 (즉시관수·관수권장 시만 표시)
    """
    if Decision["status"] not in ("즉시관수", "관수권장"):
        return
    print(f"\n  {BOLD}┌─ 관수 권고량 {'─'*45}┐{RST}")
    print(f"  │  💧 수분 부족량:  {YLW}{BOLD}{Decision['deficit_mm']:>7.2f}mm{RST}   ETc: {BOLD}{Decision['etc_mm']:>6.2f}mm{RST}            │")
    print(f"  │  🚿 권장 관수량:  {CYAN}{BOLD}{Decision['volume_l']:>7.0f}L{RST}    면적: {BOLD}{Decision['area_m2']:>6.0f}m²{RST}              │")
    print(f"  │  ⏰ 권장 시간:    이른 아침(06~08시) 또는 저녁(18~20시)                │")
    print(f"  │  💡 방식:        점적관수 권장 (증발 손실 최소화)                      │")
    print(f"  {BOLD}└{'─'*63}┘{RST}")


def Print_Advice_Panel(Advice: str, Is_AI: bool = False):
    """
    [Output] AI 또는 규칙 기반 관수 권고문 패널
    """
    Title = "🤖 Claude AI 관수 권고문" if Is_AI else "📋 관수 권고문"
    print(f"\n  {BOLD}┌─ {Title} {'─'*(52 - len(Title))}┐{RST}")
    for Raw in Advice.split("\n"):
        if not Raw.strip():
            print(f"  │{' '*63}│")
            continue
        Line = Raw
        while len(Line) > 58:
            print(f"  │  {Line[:58]}  │")
            Line = "     " + Line[58:]
        print(f"  │  {Line:<60}│")
    print(f"  {BOLD}└{'─'*63}┘{RST}")


def Print_Dashboard(Weather: dict, Soil: dict, Decision: dict,
                    Crop_Type: str, Advice: str, Is_AI: bool = False):
    """
    [Output] 전체 분석 대시보드 (모든 패널 조합)
    """
    print(f"\n{BOLD}{'='*67}{RST}")
    print(f"{BOLD}{CYAN}{'  🌾 스마트 관수 분석 대시보드':^60}{RST}")
    print(f"{BOLD}{'='*67}{RST}")
    print(f"  📅 분석일: {BOLD}{Weather.get('날짜','')}{RST}   🌾 작물: {BOLD}{Crop_Type}{RST}")

    Print_Status_Panel(Decision)
    Print_Weather_Panel(Weather)
    Print_Moisture_Gauge(Decision, Soil)
    Print_Irrigation_Panel(Decision)
    Print_Advice_Panel(Advice, Is_AI)

    print(f"\n{BOLD}{'='*67}{RST}\n")


def Print_Weekly_Chart(Weather_List: list[dict], Soil_List: list[dict]):
    """
    [Output] 주간 데이터 막대 차트
    """
    print(f"\n{BOLD}{CYAN}  ╔{'═'*66}╗{RST}")
    print(f"{BOLD}{CYAN}  ║{'📊 주간 데이터 차트 (최근 7일)':^58}║{RST}")
    print(f"{BOLD}{CYAN}  ╚{'═'*66}╝{RST}")
    print(f"  {BOLD}{'날짜':<10} {'온도':>6}  {'강수':>6}  {'토양수분':>8}   수분 게이지{RST}")
    print(f"  {'─'*72}")
    BAR_W = 22
    for W, S in zip(Weather_List, Soil_List):
        Date   = W["날짜"][5:]
        MT     = W.get("평균온도_C", 0)
        Rain   = W.get("강수량_mm", 0)
        Mois   = S.get("토양수분_Pct", 0)
        MC     = RED if Mois < 40 else YLW if Mois < 55 else GRN if Mois <= 85 else BLU
        RC     = CYAN if Rain > 0 else GRY
        Bar    = f"{MC}{'█' * int(Mois/100*BAR_W)}{'░'*(BAR_W-int(Mois/100*BAR_W))}{RST}"
        print(f"  {Date:<10} {MT:>5.1f}°C  {RC}{Rain:>6.1f}{RST}  {MC}{Mois:>6.1f}%{RST}   {Bar}")
    print(f"  {'─'*72}")
    print(f"  {RED}■{RST}<40%위험  {YLW}■{RST}40~55%주의  {GRN}■{RST}55~85%적정  {BLU}■{RST}>85%과습\n")


def Print_Quality_Report(QR: dict):
    """
    [Output] 데이터 품질 보고서
    """
    Grade  = QR.get("quality_grade", "?")
    Score  = QR.get("quality_score", 0)
    GC     = {"A": GRN, "B": YLW, "C": YLW, "D": RED}.get(Grade, RST)
    print(f"\n  {BOLD}📊 데이터 품질 보고서{RST}")
    print(f"  {'─'*50}")
    print(f"  결측값 처리:   {QR.get('total_missing', 0)}건")
    print(f"  이상값 처리:   {QR.get('total_outliers', 0)}건")
    print(f"  CSV 저장:      {GRY}{QR.get('saved_csv','')}{RST} {GRN}[LLM 분석용]{RST}")
    print(f"  Excel 저장:    {GRY}{QR.get('saved_xlsx','')}{RST} {YLW}[수치 조정용]{RST}")
    Src_Labels = {"G": "기상청API", "W": "WAMIS", "S": "시뮬레이션"}
    for Flag, Cnt in QR.get("source_counts", {}).items():
        Label = Src_Labels.get(Flag, Flag)
        print(f"  {Label:<12}: {'█'*min(Cnt,30)} ({Cnt}건)")
    print(f"  {'─'*50}")
    print(f"  품질 점수: {GC}{BOLD}{Score:.1f}/100{RST}")
    print(f"  품질 등급: {GC}{BOLD}{Grade} — {QR.get('quality_desc','')}{RST}")
    print(f"  {'─'*50}\n")


def Print_Validation_Report(MR: dict):
    """
    [Output] 모델 검증 보고서
    """
    CV   = MR.get("cv_result", {})
    FM   = MR.get("final_metrics", {})
    Hist = MR.get("train_history", [])
    Dist = MR.get("label_dist", {})

    print(f"\n  {BOLD}🔬 모델 검증 보고서{RST}")
    print(f"  {'─'*55}")
    print(f"  검증 방법:    {CV.get('method', '-')}")
    print(f"  평균 정확도:  {BOLD}{CV.get('mean_accuracy', 0):.1%}{RST}")
    print(f"  평균 F1:      {BOLD}{CV.get('mean_f1', 0):.3f}{RST}")
    print(f"  정확도 편차:  {CV.get('std_accuracy', 0):.3f}  (낮을수록 안정적)")
    print(f"  레이블 분포:  {Dist}")

    print(f"\n  {BOLD}학습 곡선{RST}")
    for Ep, Acc in enumerate(Hist[:10], 1):
        BL  = int(Acc / 100 * 25)
        BC  = GRN if Acc >= 70 else YLW if Acc >= 50 else RED
        print(f"  Epoch {Ep:>2}: {BC}{'█'*BL}{'░'*(25-BL)}{RST} {Acc:.1f}%")

    Cm = FM.get("class_metrics", {})
    if Cm:
        print(f"\n  {BOLD}클래스별 성능{RST}")
        print(f"  {'클래스':<10} {'정밀도':>8} {'재현율':>8} {'F1':>8} {'샘플':>6}")
        print(f"  {'─'*45}")
        for Label, M in Cm.items():
            FC = GRN if M["F1"] >= 0.7 else YLW if M["F1"] >= 0.5 else RED
            print(f"  {Label:<10} {M['Precision']:>8.3f} {M['Recall']:>8.3f} {FC}{M['F1']:>8.3f}{RST} {M['Support']:>6}")

    F1 = CV.get("mean_f1", 0)
    print(f"  {'─'*55}")
    if F1 >= 0.7:   Print_Msg("모델 품질 양호 (F1 >= 0.7) — 사용 가능", "success")
    elif F1 >= 0.5: Print_Msg("모델 개선 필요 — 데이터 추가 권장", "warn")
    else:           Print_Msg("모델 품질 불량 — 재학습 필요", "error")
    print()


def Print_History_Table(Records: list[dict], Last_N: int = 10):
    """
    [Output] 관수 이력 테이블
    """
    Recent = Records[-Last_N:]
    print(f"\n{BOLD}{CYAN}  ╔{'═'*68}╗{RST}")
    print(f"{BOLD}{CYAN}  ║{'📋 관수 이력 (최근 '+str(Last_N)+'건)':^58}║{RST}")
    print(f"{BOLD}{CYAN}  ╚{'═'*68}╝{RST}")
    print(f"  {BOLD}{'날짜':<12} {'작물':<12} {'수분':>7} {'판단':>8} {'관수':>6} {'실제량':>9} {'권장량':>9}{RST}")
    print(f"  {'─'*70}")
    if not Recent:
        print(f"  {GRY}  저장된 이력이 없습니다.{RST}")
    else:
        for R in Recent:
            S    = R.get("판단결과", "")
            SC   = {s: c for s, (c, _) in STATUS_CFG.items()}.get(S, GRN)
            Irr  = f"{GRN}✔완료{RST}" if R.get("관수여부") else f"{GRY}─미실시{RST}"
            Vol  = f"{CYAN}{R.get('실제관수량_L',0):>7.0f}L{RST}" if R.get("관수여부") else f"{GRY}{'─':>8}{RST}"
            print(f"  {R.get('날짜',''):<12} {R.get('작물',''):<12} {SC}{R.get('토양수분_Pct',0):>6.1f}%{RST}  {SC}{S:>8}{RST}  {Irr}  {Vol}  {GRY}{R.get('권장관수량_L',0):>7.0f}L{RST}")
    print(f"  {'─'*70}")


def Print_Stats(Stats: dict):
    """
    [Output] 관수 통계
    """
    print(f"\n{BOLD}{CYAN}  ╔{'═'*48}╗{RST}")
    print(f"{BOLD}{CYAN}  ║{'📈 관수 통계':^40}║{RST}")
    print(f"{BOLD}{CYAN}  ╚{'═'*48}╝{RST}")
    print(f"  총 기록:       {BOLD}{Stats['total']}건{RST}")
    print(f"  관수 횟수:     {GRN}{BOLD}{Stats['irrigated']}회{RST}")
    print(f"  총 관수량:     {CYAN}{BOLD}{Stats['total_vol_L']:,.0f}L{RST}")
    print(f"  절약 추정량:   {YLW}{BOLD}{Stats['saved_vol_L']:,.0f}L{RST}")
    print(f"\n  상태별 분포:")
    for S, Cnt in Stats.get("status_dist", {}).items():
        SC = {s: c for s, (c, _) in STATUS_CFG.items()}.get(S, GRY)
        print(f"    {SC}{S:<8}{RST}: {'█'*Cnt} ({Cnt}건)")
    print()


# ══════════════════════════════════════════════════════════
#  섹션 E — 전체 파이프라인 실행 (UI 오케스트레이션)
# ══════════════════════════════════════════════════════════

def Run_Pipeline(
    Crop_Type:       str,
    Season:          str,
    Area_m2:         float,
    Initial_Moisture: float,
    Days:            int,
    KMA_API_Key:     str | None,
    API_Key:         str | None,
    History_Records: list[dict],
) -> tuple:
    """
    [Input]  파이프라인 파라미터 + 이력 리스트
    [Output] (Weather_List, Soil_List, Model_Result, Decision, Advice_Result, Updated_History)
    6단계 파이프라인 진행 상황을 실시간 출력.
    """
    print(f"\n{BOLD}{CYAN}{'─'*67}{RST}")
    W_List = S_List = None
    MR = Dec = Adv = None

    # ─ 1단계: 수집 ─
    Print_Progress(1, 6, "데이터 수집")
    CW = collect_weather(Days, Season, KMA_API_Key, "data/raw")
    Print_Msg(CW["log"], "hint")
    W_Raw = CW["data"]

    # ─ 2단계: 전처리 ─
    Print_Progress(2, 6, "데이터 전처리 · 품질 관리")
    QR = preprocess(W_Raw, "data/processed")
    Print_Quality_Report(QR)
    W_List = QR["data"]

    GS = generate_soil(W_List, Crop_Type, Initial_Moisture, "data/raw")
    Soil_Raw = GS["data"]
    SQR = preprocess(Soil_Raw, "data/processed")
    S_List = SQR["data"]

    # ─ 3~4단계: 모델 ─
    Print_Progress(3, 6, "모델 개발 · 검증")
    MR = train_and_validate(W_List, S_List, Crop_Type, "data/model")
    Print_Msg(MR["log"], "hint")
    Print_Progress(4, 6, "모델 검증 결과 확인")
    Print_Validation_Report(MR)

    # ─ 5단계: 의사결정 ─
    Print_Progress(5, 6, "관수 의사결정")
    Dec = make_decision(
        MR["model"], MR["norm_min"], MR["norm_max"],
        W_List[-1], S_List[-1], W_List, Crop_Type, Area_m2
    )

    Adv = get_advice(Dec, W_List[-1], S_List[-1], Crop_Type, API_Key)

    # ─ 6단계: 출력 ─
    Print_Progress(6, 6, "결과 출력")
    Print_Dashboard(W_List[-1], S_List[-1], Dec, Crop_Type,
                    Adv["advice"], Adv["source"] == "claude_ai")

    # 관수 이력 기록
    Ans      = Safe_Input("오늘 관수를 실시하시겠습니까? (y/n)", "n").lower()
    Irrigated = Ans in ("y", "yes", "예", "ㅇ")
    Vol = 0.0
    if Irrigated:
        Vol = Safe_Float(f"실제 관수량 (L, 권장: {Dec['volume_l']:.0f}L)", Dec["volume_l"])

    Updated_History = add_history_record(
        History_Records, W_List[-1]["날짜"], Crop_Type, Dec, Irrigated, Vol
    )
    Msg = f"관수 이력 저장 완료 ({Vol:.0f}L)" if Irrigated else "미실시 이력 저장됨"
    Print_Msg(Msg, "success" if Irrigated else "hint")

    return W_List, S_List, MR, Dec, Adv, Updated_History


# ══════════════════════════════════════════════════════════
#  섹션 F — 메인 루프
# ══════════════════════════════════════════════════════════

def Main():
    os.system("cls" if os.name == "nt" else "clear")
    print(f"\n{BOLD}{GRN}")
    print("  ╔══════════════════════════════════════════════════════╗")
    print("  ║   🌾 스마트 관수 프로그램 v4.0                       ║")
    print("  ║   Frontend (UI) ←→ Backend (Logic) 분리 구조        ║")
    print("  ╚══════════════════════════════════════════════════════╝")
    print(f"{RST}")

    # ─ 초기화 ─
    History = load_history()
    API_Key     = os.environ.get("ANTHROPIC_API_KEY")
    KMA_API_Key = os.environ.get("KMA_API_KEY")

    if API_Key:     Print_Msg("Claude AI 연동 활성화됨", "success")
    if KMA_API_Key: Print_Msg("기상청 API 활성화됨", "success")
    if not KMA_API_Key:
        Print_Msg("기상청 API 키 없음 → WAMIS 크롤링 또는 시뮬레이션 사용", "hint")
        Print_Msg("설정: set KMA_API_KEY=<키>  (data.go.kr 공공데이터포털)", "hint")

    CROP_LIST    = list(CROP_THRESHOLDS.keys())
    SEASON_LIST  = ["봄", "여름", "가을", "겨울"]
    Month        = datetime.date.today().month
    DEF_SEASON   = 0 if Month in (3,4,5) else 1 if Month in (6,7,8) else 2 if Month in (9,10,11) else 3

    Cached_W: list[dict] = []
    Cached_S: list[dict] = []

    while True:
        print(f"\n{BOLD}  ┌─ 메인 메뉴 {'─'*42}┐{RST}")
        print(f"  │  1. 🚀 전체 파이프라인 실행 (수집→전처리→모델→결과)  │")
        print(f"  │  2. 📊 주간 데이터 차트                              │")
        print(f"  │  3. 📋 관수 이력 · 통계                              │")
        print(f"  │  4. 💾 데이터 내보내기 (CSV·LLM용 / Excel·조정용)   │")
        print(f"  │  5. 🔑 API 키 설정                                   │")
        print(f"  │  7. ❓ 사용법 · 구조 안내                            │")
        print(f"  │  0. 🚪 종료                                         │")
        print(f"{BOLD}  └{'─'*54}┘{RST}")

        Choice = Safe_Input("선택")

        if Choice == "1":
            Crop   = Select_From_List("작물 선택",  CROP_LIST,   6)
            Season = Select_From_List("계절 선택",  SEASON_LIST, DEF_SEASON)
            Days   = max(5, int(Safe_Float("수집 일수 (최소 10 권장)", 30)))
            Area   = Safe_Float("재배 면적 (m²)", 1000.0)
            Mois   = Safe_Float("현재 토양 수분 추정 (%)", 65.0)

            Cached_W, Cached_S, MR, Dec, Adv, History = Run_Pipeline(
                Crop, Season, Area, Mois, Days, KMA_API_Key, API_Key, History
            )

        elif Choice == "2":
            if Cached_W and Cached_S:
                Print_Weekly_Chart(Cached_W[-7:], Cached_S[-7:])
            else:
                Print_Msg("먼저 메뉴 1에서 파이프라인을 실행해주세요.", "warn")

        elif Choice == "3":
            Print_History_Table(History, 10)
            Print_Stats(calc_stats(History))

        elif Choice == "4":
            if not Cached_W:
                Print_Msg("먼저 메뉴 1에서 파이프라인을 실행해주세요.", "warn")
                continue
            Print_Section("데이터 내보내기")
            Print_Msg("CSV  → LLM(Claude) 분석에 최적화된 텍스트 형식", "hint")
            Print_Msg("Excel → 수치 조정 및 시각적 검토에 편리한 셀 형식", "hint")
            Name = Safe_Input("파일 이름 기반 (예: farm_20260101)", "export")
            ER   = export_data(Cached_W, Cached_S, History, Name)
            Print_Msg(ER["log"], "success")
            for P in ER["csv_paths"]:  Print_Msg(f"CSV:   {P}", "hint")
            for P in ER["xlsx_paths"]: Print_Msg(f"Excel: {P}", "hint")

        elif Choice == "5":
            Print_Section("API 키 설정")
            print(f"  1. Claude AI 키  — https://console.anthropic.com")
            print(f"  2. 기상청 API 키 — https://data.go.kr")
            Sub = Safe_Input("선택 (1/2)")
            if Sub == "1":
                K = input("  Claude API 키: ").strip()
                if K: API_Key = K; os.environ["ANTHROPIC_API_KEY"] = K; Print_Msg("Claude AI 키 설정 완료", "success")
            elif Sub == "2":
                K = input("  기상청 API 키: ").strip()
                if K: KMA_API_Key = K; os.environ["KMA_API_KEY"] = K; Print_Msg("기상청 키 설정 완료", "success")

        elif Choice == "7":
            print(f"""
  {BOLD}[ 📖 프로그램 구조 안내 ]{RST}

  {BOLD}Frontend / Backend 분리 구조{RST}
  ┌──────────────────────────────────────────────────┐
  │  ui_frontend.py  —  FRONTEND (화면 · 입출력)     │
  │  → 사용자 입력 받기 (Input)                      │
  │  → 분석 결과 화면에 출력 (Output)                │
  │  → backend_logic.py 함수 호출만 함               │
  └────────────────────┬─────────────────────────────┘
                       │ 함수 호출 / 결과 반환
  ┌────────────────────▼─────────────────────────────┐
  │  backend_logic.py  —  BACKEND (연산 · 저장)      │
  │  → 기상 수집, 전처리, 모델 학습, 의사결정        │
  │  → CSV · Excel · JSON 저장                       │
  │  → print() 없음 — 결과는 dict/list 로 반환      │
  └──────────────────────────────────────────────────┘

  {BOLD}데이터 출처 (결과 보고서 기재 필요){RST}
  • 기상청 기상자료개방포털: https://data.kma.go.kr
  • WAMIS 한국수자원공사:   https://www.wamis.go.kr
  • 시뮬레이션: 계절별 통계 파라미터 기반

  {BOLD}파일 형식 선택 기준{RST}
  • CSV   → LLM(Claude) 분석 시 최적 (텍스트 기반)
  • Excel → 수치 직접 조정, 셀 서식 활용 시 편리

  {BOLD}스마트폰에서 사용하려면?{RST}
  → 아래 방법들 중 선택 가능합니다.
  1. Termux (Android): 안드로이드 앱으로 Python 실행
  2. iSH  (iPhone):   iOS 앱으로 리눅스 쉘 + Python
  3. Streamlit 웹앱:  python -m streamlit run app.py
     → 브라우저로 스마트폰에서 접속 가능
  4. FastAPI + HTML:  REST API 서버 + 모바일 웹 UI
  자세한 내용은 메인 메뉴 종료 후 README.md 참조
            """)

        elif Choice == "0":
            print(f"\n  {GRN}종료합니다. 🌾 좋은 농사 되세요!{RST}\n")
            sys.exit(0)
        else:
            Print_Msg("0~7 사이에서 선택해주세요.", "error")


if __name__ == "__main__":
    try:
        Main()
    except KeyboardInterrupt:
        print(f"\n\n  {YLW}프로그램 중단 (Ctrl+C){RST}\n")
        sys.exit(0)
    except Exception as E:
        print(f"\n  {RED}[예기치 않은 오류] {E}{RST}")
        import traceback; traceback.print_exc()
        sys.exit(1)

"""
==============================================================
  backend_logic.py  —  BACKEND (Logic Layer)
==============================================================

  [역할]
  "내부에서 돌아가는 코딩" — 사용자에게 보이지 않는 연산 담당.
  화면 출력이 단 한 줄도 없습니다.
  UI(Frontend)는 이 파일의 함수를 호출하기만 합니다.

  [포함 기능]
  ① 기상 데이터 수집 (API / WAMIS 크롤링 / 시뮬레이션)
  ② 토양 데이터 생성 (물 수지 모델)
  ③ 데이터 전처리 · 품질 관리
  ④ 관수 예측 모델 개발 · K-Fold 검증
  ⑤ 관수 의사결정 (모델 기반)
  ⑥ AI 권고문 생성 (Claude API / 규칙 기반)
  ⑦ CSV · Excel · JSON 저장 · 불러오기

  [Input / Output 규약]
  - 모든 public 함수는 순수 Python 자료형(dict, list, str, float)만 반환
  - 터미널 출력(print) 금지 — 로그는 Return 값에 포함
  - 예외는 모두 try/except로 포착하여 {"error": "..."} 형태로 반환
==============================================================
"""

import csv
import json
import math
import os
import random
import datetime
import statistics
from collections import Counter

# ── 선택적 패키지 (없어도 실행되도록 guard) ──────────────
try:
    import requests
    from bs4 import BeautifulSoup
    _REQUESTS_OK = True
except ImportError:
    _REQUESTS_OK = False

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    _EXCEL_OK = True
except ImportError:
    _EXCEL_OK = False

try:
    import pandas as pd
    _PANDAS_OK = True
except ImportError:
    _PANDAS_OK = False


# ══════════════════════════════════════════════════════════
#  섹션 A — 상수 · 도메인 지식
# ══════════════════════════════════════════════════════════

CROP_THRESHOLDS: dict = {
    "벼":            {"Min": 60, "Opt": 80, "Max": 100},
    "밀/보리":       {"Min": 40, "Opt": 60, "Max": 75},
    "옥수수":        {"Min": 45, "Opt": 65, "Max": 80},
    "콩":            {"Min": 40, "Opt": 58, "Max": 70},
    "감자":          {"Min": 50, "Opt": 70, "Max": 85},
    "토마토":        {"Min": 55, "Opt": 72, "Max": 85},
    "상추/채소":     {"Min": 50, "Opt": 68, "Max": 80},
    "과수(사과/배)": {"Min": 40, "Opt": 60, "Max": 75},
}

CROP_KC: dict = {
    "벼": 1.20, "밀/보리": 0.85, "옥수수": 1.15, "콩": 1.05,
    "감자": 1.10, "토마토": 1.15, "상추/채소": 1.00, "과수(사과/배)": 0.90,
}

SEASON_PARAMS: dict = {
    "봄":   {"Temp_Base": 15, "Temp_Var": 8,  "Humid_Base": 60, "Rain_Prob": 0.30},
    "여름": {"Temp_Base": 30, "Temp_Var": 5,  "Humid_Base": 78, "Rain_Prob": 0.45},
    "가을": {"Temp_Base": 18, "Temp_Var": 7,  "Humid_Base": 55, "Rain_Prob": 0.25},
    "겨울": {"Temp_Base": 2,  "Temp_Var": 8,  "Humid_Base": 45, "Rain_Prob": 0.20},
}

VALID_RANGE: dict = {
    "평균온도_C":      (-40.0, 50.0),
    "상대습도_Pct":    (0.0, 100.0),
    "강수량_mm":       (0.0, 300.0),
    "풍속_ms":         (0.0, 50.0),
    "일사량_MJ":       (0.0, 35.0),
    "증발산량ET0_mm":  (0.0, 15.0),
}

IRRIGATION_LABELS: dict = {0: "즉시관수", 1: "관수권장", 2: "적정상태", 3: "과습주의"}
FEATURE_COLS = [
    "토양수분_Pct", "평균온도_C", "상대습도_Pct",
    "강수량_mm", "증발산량ET0_mm", "풍속_ms",
    "최근3일강수_mm", "이동평균수분_Pct",
]


# ══════════════════════════════════════════════════════════
#  섹션 B — 데이터 수집
# ══════════════════════════════════════════════════════════

def _calc_et0(Mean_Temp: float, Temp_Var: float, Solar_MJ: float) -> float:
    """ET₀ 계산 (Penman-Monteith 간이식). 내부 전용."""
    try:
        return round(max(0.0, 0.0023 * (Mean_Temp + 17.8) * math.sqrt(max(0, Temp_Var)) * Solar_MJ * 0.1), 2)
    except Exception:
        return 0.5


def _sim_weather(Days: int, Season: str, Start_Date: datetime.date) -> list[dict]:
    """시뮬레이션 날씨 생성. 내부 전용."""
    P = SEASON_PARAMS.get(Season, SEASON_PARAMS["여름"])
    Result = []
    for I in range(Days):
        D     = Start_Date + datetime.timedelta(days=I)
        MT    = round(P["Temp_Base"] + random.uniform(-P["Temp_Var"], P["Temp_Var"]), 1)
        Rain  = round(random.uniform(3, 35), 1) if random.random() < P["Rain_Prob"] else 0.0
        Solar = round(random.uniform(8, 24), 1)
        Result.append({
            "날짜":            str(D),
            "출처":            "시뮬레이션",
            "최고온도_C":      round(MT + random.uniform(2, 5), 1),
            "최저온도_C":      round(MT - random.uniform(2, 5), 1),
            "평균온도_C":      MT,
            "상대습도_Pct":    round(min(100, max(10, P["Humid_Base"] + random.uniform(-20, 20))), 1),
            "강수량_mm":       Rain,
            "풍속_ms":         round(random.uniform(0.5, 7), 1),
            "일사량_MJ":       Solar,
            "증발산량ET0_mm":  _calc_et0(MT, P["Temp_Var"], Solar),
            "데이터품질_Flag": "S",
        })
    return Result


def collect_weather(
    Days: int = 30,
    Season: str = "여름",
    KMA_API_Key: str | None = None,
    Save_Dir: str = "data/raw",
) -> dict:
    """
    [INPUT]  Days, Season, KMA_API_Key, Save_Dir
    [OUTPUT] {"data": [...], "source": str, "saved_path": str, "log": str}

    기상청 API → WAMIS 크롤링 → 시뮬레이션 순으로 시도.
    데이터 출처:
      - 기상청 기상자료개방포털 https://data.kma.go.kr
      - WAMIS 한국수자원공사   https://www.wamis.go.kr
    """
    Today      = datetime.date.today()
    Start_Date = Today - datetime.timedelta(days=Days - 1)
    Log_Lines  = []
    Data: list[dict] | None = None
    Source = "시뮬레이션"

    # ◆ 기상청 API 시도
    if KMA_API_Key and _REQUESTS_OK:
        try:
            Params = {
                "serviceKey": KMA_API_Key, "pageNo": "1", "numOfRows": "100",
                "dataType": "JSON", "dataCd": "ASOS", "dateCd": "DAY",
                "startDt": Start_Date.strftime("%Y%m%d"),
                "endDt":   Today.strftime("%Y%m%d"), "stnIds": "131",
            }
            R = requests.get(
                "http://apis.data.go.kr/1360000/AsosDalyInfoService/getWthrDataList",
                params=Params, timeout=10
            )
            Items = R.json().get("response", {}).get("body", {}).get("items", {}).get("item", [])
            if Items:
                Data = []
                for Item in Items:
                    MT = float(Item.get("avgTa", 15) or 15)
                    Sol = float(Item.get("sumGsr", 12) or 12)
                    Data.append({
                        "날짜": Item.get("tm", ""), "출처": "기상청API",
                        "평균온도_C": MT,
                        "최고온도_C": float(Item.get("maxTa", MT+3) or MT+3),
                        "최저온도_C": float(Item.get("minTa", MT-3) or MT-3),
                        "상대습도_Pct": float(Item.get("avgRhm", 60) or 60),
                        "강수량_mm":   float(Item.get("sumRn", 0) or 0),
                        "풍속_ms":     float(Item.get("avgWs", 1.5) or 1.5),
                        "일사량_MJ":   Sol,
                        "증발산량ET0_mm": _calc_et0(MT, 6, Sol),
                        "데이터품질_Flag": "G",
                    })
                Source = "기상청API"
                Log_Lines.append(f"기상청 API 수집 성공: {len(Data)}건")
        except Exception as E:
            Log_Lines.append(f"기상청 API 실패: {E}")

    # ◆ WAMIS 크롤링 시도
    if Data is None and _REQUESTS_OK:
        try:
            Headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120"}
            R = requests.get(
                "https://www.wamis.go.kr/wkw/wk_flw_slct.aspx",
                params={"obscd": "131",
                        "startdt": Start_Date.strftime("%Y%m%d"),
                        "enddt":   Today.strftime("%Y%m%d"),
                        "gubun": "D"},
                headers=Headers, timeout=10
            )
            Soup  = BeautifulSoup(R.text, "html.parser")
            Table = Soup.find("table")
            if Table:
                Rows = Table.find_all("tr")
                Tmp  = []
                for Row in Rows[2:]:
                    Cells = Row.find_all("td")
                    if len(Cells) < 4:
                        continue
                    MT  = float(Cells[2].get_text(strip=True) or 15)
                    Sol = float(Cells[5].get_text(strip=True) or 12) if len(Cells) > 5 else 12.0
                    Tmp.append({
                        "날짜": Cells[0].get_text(strip=True), "출처": "WAMIS크롤링",
                        "평균온도_C": MT,
                        "최고온도_C": MT + 3,
                        "최저온도_C": MT - 3,
                        "상대습도_Pct": float(Cells[3].get_text(strip=True) or 60),
                        "강수량_mm":   float(Cells[1].get_text(strip=True) or 0),
                        "풍속_ms":     float(Cells[4].get_text(strip=True) or 1.5) if len(Cells) > 4 else 1.5,
                        "일사량_MJ":   Sol,
                        "증발산량ET0_mm": _calc_et0(MT, 6, Sol),
                        "데이터품질_Flag": "W",
                    })
                if Tmp:
                    Data   = Tmp
                    Source = "WAMIS크롤링"
                    Log_Lines.append(f"WAMIS 크롤링 성공: {len(Data)}건")
        except Exception as E:
            Log_Lines.append(f"WAMIS 크롤링 실패: {E}")

    # ◆ 시뮬레이션 Fallback
    if Data is None:
        Data = _sim_weather(Days, Season, Start_Date)
        Log_Lines.append(f"시뮬레이션 생성: {len(Data)}건 (Season={Season})")

    # CSV 저장
    Saved = ""
    try:
        os.makedirs(Save_Dir, exist_ok=True)
        Ts   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        Path = os.path.join(Save_Dir, f"weather_raw_{Ts}.csv")
        Keys = list(Data[0].keys()) if Data else []
        with open(Path, "w", newline="", encoding="utf-8-sig") as F:
            W = csv.DictWriter(F, fieldnames=Keys)
            W.writeheader(); W.writerows(Data)
        Saved = Path
        Log_Lines.append(f"Raw CSV 저장: {Path}")
    except Exception as E:
        Log_Lines.append(f"CSV 저장 오류: {E}")

    return {"data": Data, "source": Source, "saved_path": Saved, "log": " | ".join(Log_Lines)}


# ══════════════════════════════════════════════════════════
#  섹션 C — 토양 데이터 생성
# ══════════════════════════════════════════════════════════

def generate_soil(
    Weather_List: list[dict],
    Crop_Type: str,
    Initial_Moisture: float = 65.0,
    Save_Dir: str = "data/raw",
) -> dict:
    """
    [INPUT]  Weather_List, Crop_Type, Initial_Moisture, Save_Dir
    [OUTPUT] {"data": [...], "saved_path": str}

    물 수지 모델로 토양 수분 시뮬레이션.
    """
    Kc       = CROP_KC.get(Crop_Type, 1.0)
    Moisture = Initial_Moisture
    Result   = []
    for W in Weather_List:
        Rain_E   = W.get("강수량_mm", 0) * 0.6
        ET_E     = W.get("증발산량ET0_mm", 0) * Kc * 3.0
        Drain    = max(0.0, (Moisture - 90) * 0.3)
        Moisture = max(5.0, min(100.0, Moisture + Rain_E - ET_E - Drain + random.uniform(-1.5, 1.5)))
        EC_Base  = 0.8 + (Moisture / 100) * 1.5
        Result.append({
            "날짜":          W.get("날짜", ""),
            "토양수분_Pct":  round(Moisture, 1),
            "토양온도_C":    round(W.get("평균온도_C", 15) * 0.85 + random.uniform(-1, 1), 1),
            "전기전도도_dSm": round(min(4.0, max(0.1, EC_Base + random.uniform(-0.3, 0.3))), 2),
            "pH":            round(random.uniform(5.5, 7.5), 1),
        })

    Saved = ""
    try:
        os.makedirs(Save_Dir, exist_ok=True)
        Ts   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        Path = os.path.join(Save_Dir, f"soil_raw_{Ts}.csv")
        with open(Path, "w", newline="", encoding="utf-8-sig") as F:
            W2 = csv.DictWriter(F, fieldnames=list(Result[0].keys()))
            W2.writeheader(); W2.writerows(Result)
        Saved = Path
    except Exception:
        pass

    return {"data": Result, "saved_path": Saved}


# ══════════════════════════════════════════════════════════
#  섹션 D — 데이터 전처리 · 품질 관리
# ══════════════════════════════════════════════════════════

def _impute(Data: list[dict], Col: str, Idx: int) -> float:
    """결측값 선형 보간. 내부 전용."""
    Prev = Next = None
    for J in range(Idx - 1, -1, -1):
        V = Data[J].get(Col)
        if V not in (None, ""):
            try: Prev = float(V); break
            except (ValueError, TypeError): pass
    for J in range(Idx + 1, len(Data)):
        V = Data[J].get(Col)
        if V not in (None, ""):
            try: Next = float(V); break
            except (ValueError, TypeError): pass
    if Prev is not None and Next is not None:
        return round((Prev + Next) / 2, 2)
    if Prev is not None: return Prev
    if Next is not None: return Next
    Vals = [float(R[Col]) for R in Data if R.get(Col) not in (None, "")]
    return round(statistics.mean(Vals), 2) if Vals else 0.0


def preprocess(
    Raw_Data: list[dict],
    Save_Dir: str = "data/processed",
) -> dict:
    """
    [INPUT]  Raw_Data (기상 또는 토양 리스트), Save_Dir
    [OUTPUT] {
        "data":           전처리된 리스트,
        "quality_score":  float (0~100),
        "quality_grade":  str (A/B/C/D),
        "quality_desc":   str,
        "total_missing":  int,
        "total_outliers": int,
        "source_counts":  dict,
        "saved_csv":      str,
        "saved_xlsx":     str,
        "log":            str,
    }

    1. 결측값 탐지 → 선형 보간
    2. 이상값 탐지 → 범위 클리핑
    3. 논리 검증 (최고 >= 평균 >= 최저)
    4. 품질 점수 A/B/C/D
    5. CSV (LLM용) + Excel (수치조정용) 저장
    """
    Num_Cols      = list(VALID_RANGE.keys())
    Processed     = []
    Total_Missing = 0
    Total_Out     = 0
    Source_Counts: dict = {}

    for Idx, Row in enumerate(Raw_Data):
        Source_Counts[Row.get("데이터품질_Flag", "S")] = \
            Source_Counts.get(Row.get("데이터품질_Flag", "S"), 0) + 1
        C = dict(Row)
        Miss_Cnt = Out_Cnt = 0

        # 결측값 처리
        for Col in Num_Cols:
            if C.get(Col) in (None, ""):
                C[Col] = _impute(Raw_Data, Col, Idx); Miss_Cnt += 1

        # 이상값 클리핑
        for Col, (Lo, Hi) in VALID_RANGE.items():
            try:
                V = float(C.get(Col, 0) or 0)
                if V < Lo or V > Hi:
                    C[Col] = round(max(Lo, min(Hi, V)), 2); Out_Cnt += 1
            except (ValueError, TypeError):
                C[Col] = Lo; Out_Cnt += 1

        # 기온 논리 검증
        try:
            MX = float(C.get("최고온도_C", 0) or 0)
            MT = float(C.get("평균온도_C", 0) or 0)
            MN = float(C.get("최저온도_C", 0) or 0)
            if not (MX >= MT >= MN):
                C["최고온도_C"] = round(MT + abs(MX - MT), 1)
                C["최저온도_C"] = round(MT - abs(MT - MN), 1)
                Out_Cnt += 1
        except (ValueError, TypeError):
            pass

        C["전처리_Flag"]      = "P" if (Miss_Cnt + Out_Cnt) > 0 else "OK"
        C["결측처리_Count"]   = Miss_Cnt
        C["이상값처리_Count"] = Out_Cnt
        Processed.append(C)
        Total_Missing += Miss_Cnt; Total_Out += Out_Cnt

    # 품질 점수 산출
    N = len(Processed) or 1
    Score = 100.0
    Score -= (Total_Missing / (N * len(Num_Cols))) * 30
    Score -= (Total_Out     / (N * len(Num_Cols))) * 20
    Score -= (Source_Counts.get("S", 0) / N) * 20
    Score  = round(max(0.0, min(100.0, Score)), 1)
    Grade  = "D"
    for G, T in [("A", 90), ("B", 70), ("C", 50), ("D", 0)]:
        if Score >= T: Grade = G; break
    Desc   = {"A": "우수 — 바로 사용 가능", "B": "양호 — 일부 주의",
               "C": "보통 — 추가 수집 권장", "D": "불량 — 재수집 필요"}.get(Grade, "")

    # 저장
    Saved_CSV = Saved_XLSX = ""
    try:
        os.makedirs(Save_Dir, exist_ok=True)
        Ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        # CSV — LLM 분석용
        Saved_CSV = os.path.join(Save_Dir, f"processed_{Ts}.csv")
        Keys = list(Processed[0].keys()) if Processed else []
        with open(Saved_CSV, "w", newline="", encoding="utf-8-sig") as F:
            W = csv.DictWriter(F, fieldnames=Keys, extrasaction="ignore")
            W.writeheader(); W.writerows(Processed)
        # Excel — 수치 조정용
        if _EXCEL_OK:
            Saved_XLSX = os.path.join(Save_Dir, f"processed_{Ts}.xlsx")
            WB = openpyxl.Workbook()
            WS = WB.active
            WS.title = "전처리데이터"
            WS.append(Keys)
            for R in Processed:
                WS.append([R.get(K, "") for K in Keys])
            WB.save(Saved_XLSX)
    except Exception as E:
        Saved_CSV = Saved_XLSX = f"저장오류: {E}"

    return {
        "data":           Processed,
        "quality_score":  Score,
        "quality_grade":  Grade,
        "quality_desc":   Desc,
        "total_missing":  Total_Missing,
        "total_outliers": Total_Out,
        "source_counts":  Source_Counts,
        "saved_csv":      Saved_CSV,
        "saved_xlsx":     Saved_XLSX,
        "log":            f"전처리 완료 {N}행 | 품질 {Grade} ({Score}점)",
    }


# ══════════════════════════════════════════════════════════
#  섹션 E — 모델 개발 · 검증
# ══════════════════════════════════════════════════════════

class _RuleWeightedModel:
    """가중치 규칙 기반 관수 예측 모델. 내부 전용."""
    def __init__(self, Crop_Type: str):
        self.Crop_Type   = Crop_Type
        self.Weights     = [0.40, 0.15, 0.10, 0.15, 0.10, 0.05, 0.03, 0.02]
        self.Train_Hist  = []

    def _predict_one(self, FV: list[float], NMin: dict, NMax: dict) -> int:
        T = CROP_THRESHOLDS.get(self.Crop_Type, CROP_THRESHOLDS["상추/채소"])
        S_Range = NMax.get("토양수분_Pct", 100) - NMin.get("토양수분_Pct", 0)
        Soil    = FV[0] * S_Range + NMin.get("토양수분_Pct", 0)
        R_Range = NMax.get("강수량_mm", 100)   - NMin.get("강수량_mm", 0)
        Rain    = FV[3] * R_Range
        if Soil < T["Min"]:               return 0
        elif Soil < T["Opt"] * 0.85:      return 2 if Rain > 20 else 1
        elif Soil > T["Max"]:             return 3
        else:                              return 2

    def train(self, FM: list, Labels: list, NMin: dict, NMax: dict, Epochs: int = 10):
        LR = 0.01
        for _ in range(Epochs):
            Correct = 0
            for FV, TL in zip(FM, Labels):
                PL = self._predict_one(FV, NMin, NMax)
                if PL == TL: Correct += 1
                else:
                    Err = float(TL - PL) * LR
                    self.Weights = [W + Err * V for W, V in zip(self.Weights, FV)]
                    TW = sum(abs(W) for W in self.Weights)
                    if TW > 0: self.Weights = [W / TW for W in self.Weights]
            self.Train_Hist.append(round(Correct / max(len(Labels), 1) * 100, 1))

    def predict_batch(self, FM: list, NMin: dict, NMax: dict) -> list[int]:
        return [self._predict_one(FV, NMin, NMax) for FV in FM]


def _norm(Data: list[dict]) -> tuple[list, dict, dict]:
    """Min-Max 정규화. 내부 전용."""
    NMin = {}; NMax = {}
    for Col in FEATURE_COLS:
        Vals = [float(R[Col]) for R in Data if R.get(Col) not in (None, "")]
        NMin[Col] = min(Vals) if Vals else 0.0
        NMax[Col] = max(Vals) if Vals else 1.0
        if NMax[Col] == NMin[Col]: NMax[Col] = NMin[Col] + 1.0
    FM = []
    for R in Data:
        Vec = []
        for Col in FEATURE_COLS:
            try: V = float(R.get(Col) or NMin[Col])
            except (ValueError, TypeError): V = NMin[Col]
            Vec.append(round((V - NMin[Col]) / (NMax[Col] - NMin[Col]), 4))
        FM.append(Vec)
    return FM, NMin, NMax


def _labels(Data: list[dict], Crop_Type: str) -> list[int]:
    """레이블 생성. 내부 전용."""
    T = CROP_THRESHOLDS.get(Crop_Type, CROP_THRESHOLDS["상추/채소"])
    Out = []
    for R in Data:
        M = float(R.get("토양수분_Pct", 65) or 65)
        Rain = float(R.get("최근3일강수_mm", 0) or 0)
        if M < T["Min"]:            Out.append(0)
        elif M < T["Opt"] * 0.85:  Out.append(2 if Rain > 20 else 1)
        elif M > T["Max"]:          Out.append(3)
        else:                        Out.append(2)
    return Out


def _augment(Data: list[dict]) -> list[dict]:
    """파생 피처 추가. 내부 전용."""
    Aug = []
    for I, R in enumerate(Data):
        NR = dict(R)
        NR["최근3일강수_mm"]   = round(sum(float(Data[J].get("강수량_mm", 0) or 0) for J in range(max(0, I-2), I+1)), 1)
        Window = [float(Data[J].get("토양수분_Pct", 65) or 65) for J in range(max(0, I-2), I+1)]
        NR["이동평균수분_Pct"] = round(statistics.mean(Window), 1)
        Aug.append(NR)
    return Aug


def _calc_metrics(True_L: list[int], Pred_L: list[int]) -> dict:
    """분류 성능 지표 계산. 내부 전용."""
    N = 4
    CM = [[0]*N for _ in range(N)]
    for T, P in zip(True_L, Pred_L):
        if 0 <= T < N and 0 <= P < N: CM[T][P] += 1
    Acc = sum(CM[I][I] for I in range(N)) / max(len(True_L), 1)
    Class_M = {}
    for I in range(N):
        TP = CM[I][I]
        FP = sum(CM[J][I] for J in range(N)) - TP
        FN = sum(CM[I][J] for J in range(N)) - TP
        Prec = TP / max(TP + FP, 1)
        Rec  = TP / max(TP + FN, 1)
        F1   = 2 * Prec * Rec / max(Prec + Rec, 1e-9)
        Class_M[IRRIGATION_LABELS[I]] = {
            "Precision": round(Prec, 3), "Recall": round(Rec, 3),
            "F1": round(F1, 3), "Support": sum(CM[I]),
        }
    Macro_F1 = statistics.mean(V["F1"] for V in Class_M.values())
    return {"accuracy": round(Acc, 3), "macro_f1": round(Macro_F1, 3),
            "class_metrics": Class_M, "confusion_matrix": CM}


def train_and_validate(
    Weather_Data: list[dict],
    Soil_Data: list[dict],
    Crop_Type: str,
    Save_Dir: str = "data/model",
) -> dict:
    """
    [INPUT]  Weather_Data, Soil_Data, Crop_Type, Save_Dir
    [OUTPUT] {
        "model":        _RuleWeightedModel 객체 (Frontend로 전달),
        "norm_min":     dict,
        "norm_max":     dict,
        "cv_result":    dict (교차검증 결과),
        "final_metrics":dict (최종 성능),
        "train_history":list[float],
        "label_dist":   dict,
        "log":          str,
    }

    파이프라인:
    ① 병합 → ② 파생 피처 → ③ 레이블 → ④ 정규화
    → ⑤ Train/Test 분할 → ⑥ 학습 → ⑦ K-Fold 검증
    → ⑧ 최종 평가 → ⑨ 모델 저장
    """
    # ① 병합
    Merged = []
    Soil_Map = {S["날짜"]: S for S in Soil_Data}
    for W in Weather_Data:
        S = Soil_Map.get(W.get("날짜", ""), {})
        if not S and len(Soil_Data) == len(Weather_Data):
            S = Soil_Data[len(Merged)]
        Merged.append({**W, **S})

    # ② 파생 피처
    Aug    = _augment(Merged)
    Labels = _labels(Aug, Crop_Type)
    Dist   = {IRRIGATION_LABELS[K]: V for K, V in Counter(Labels).items()}

    # ④ 정규화
    FM, NMin, NMax = _norm(Aug)

    # ⑤ Train/Test 분할 (80:20)
    Split = max(1, int(len(Aug) * 0.8))
    TrF, TrL = FM[:Split], Labels[:Split]
    TeF, TeL = FM[Split:], Labels[Split:]

    # ⑥ 학습
    Model = _RuleWeightedModel(Crop_Type)
    if TrF: Model.train(TrF, TrL, NMin, NMax)

    # ⑦ K-Fold 교차 검증 (K=3)
    K    = 3
    N    = len(Aug)
    Idxs = list(range(N)); random.shuffle(Idxs)
    Fold_Accs = []; Fold_F1s = []
    for Ki in range(K):
        VStart = Ki * (N // K)
        VEnd   = VStart + (N // K) if Ki < K - 1 else N
        Val_Idx  = Idxs[VStart:VEnd]
        Tr_Idx   = [I for I in Idxs if I not in set(Val_Idx)]
        FM_fold  = _RuleWeightedModel(Crop_Type)
        if Tr_Idx: FM_fold.train([FM[I] for I in Tr_Idx], [Labels[I] for I in Tr_Idx], NMin, NMax, Epochs=3)
        Preds  = FM_fold.predict_batch([FM[I] for I in Val_Idx], NMin, NMax)
        M      = _calc_metrics([Labels[I] for I in Val_Idx], Preds)
        Fold_Accs.append(M["accuracy"]); Fold_F1s.append(M["macro_f1"])

    CV = {
        "mean_accuracy": round(statistics.mean(Fold_Accs), 3),
        "mean_f1":       round(statistics.mean(Fold_F1s), 3),
        "std_accuracy":  round(statistics.stdev(Fold_Accs) if len(Fold_Accs) > 1 else 0.0, 3),
        "method":        f"{K}-Fold 교차검증",
    }

    # ⑧ 최종 평가
    Final_P = Model.predict_batch(TeF if TeF else FM, NMin, NMax)
    Final_M = _calc_metrics(TeL if TeL else Labels, Final_P)

    # ⑨ 저장
    try:
        os.makedirs(Save_Dir, exist_ok=True)
        Ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        with open(os.path.join(Save_Dir, f"weights_{Ts}.csv"), "w", newline="", encoding="utf-8-sig") as F:
            W = csv.writer(F)
            W.writerow(["피처", "가중치"])
            for Col, Wt in zip(FEATURE_COLS, Model.Weights): W.writerow([Col, round(Wt, 6)])
        with open(os.path.join(Save_Dir, f"cv_report_{Ts}.json"), "w", encoding="utf-8") as F:
            json.dump({"cv": CV, "final": Final_M, "crop": Crop_Type, "ts": Ts}, F, ensure_ascii=False, indent=2)
    except Exception as E:
        pass

    return {
        "model":         Model,
        "norm_min":      NMin,
        "norm_max":      NMax,
        "cv_result":     CV,
        "final_metrics": Final_M,
        "train_history": Model.Train_Hist,
        "label_dist":    Dist,
        "log":           f"모델 학습 완료 | CV F1={CV['mean_f1']:.3f} | 정확도={Final_M['accuracy']:.1%}",
    }


# ══════════════════════════════════════════════════════════
#  섹션 F — 관수 의사결정
# ══════════════════════════════════════════════════════════

def make_decision(
    Model:         "_RuleWeightedModel",
    Norm_Min:      dict,
    Norm_Max:      dict,
    Weather_Today: dict,
    Soil_Today:    dict,
    Weather_Hist:  list[dict],
    Crop_Type:     str,
    Area_m2:       float,
) -> dict:
    """
    [INPUT]  모델·정규화값·오늘 기상·토양·이력·작물·면적
    [OUTPUT] {
        "status":          str  (즉시관수/관수권장/적정상태/과습주의),
        "urgency":         int  (0~5),
        "reason":          str,
        "soil_moisture":   float,
        "threshold_min":   float,
        "threshold_opt":   float,
        "threshold_max":   float,
        "etc_mm":          float,
        "deficit_mm":      float,
        "volume_l":        float,
        "recent_rain_mm":  float,
        "heat_stress":     bool,
        "area_m2":         float,
        "label_int":       int,
    }
    """
    T    = CROP_THRESHOLDS.get(Crop_Type, CROP_THRESHOLDS["상추/채소"])
    Kc   = CROP_KC.get(Crop_Type, 1.0)
    Rain3 = sum(W.get("강수량_mm", 0) for W in Weather_Hist[-3:])
    Mois  = Soil_Today.get("토양수분_Pct", 65)

    Single = {**Weather_Today, **Soil_Today,
              "최근3일강수_mm": Rain3,
              "이동평균수분_Pct": Mois}
    FM, _, _ = _norm([Single])
    Label = Model._predict_one(FM[0], Norm_Min, Norm_Max) if FM else 2
    Status = IRRIGATION_LABELS.get(Label, "적정상태")

    ET0   = Weather_Today.get("증발산량ET0_mm", 0)
    ETC   = ET0 * Kc
    Eff_R = min(Weather_Today.get("강수량_mm", 0) * 0.8, ETC)
    Def   = max(0.0, ETC - Eff_R)
    Vol   = round(Def * Area_m2, 1)

    Reasons = {
        "즉시관수": f"토양 수분 {Mois}%가 최소 기준({T['Min']}%) 미달",
        "관수권장": f"토양 수분 {Mois}%가 최적({T['Opt']}%) 이하",
        "적정상태": f"토양 수분 {Mois}%가 적정 범위 유지 중",
        "과습주의": f"토양 수분 {Mois}%가 최대({T['Max']}%) 초과",
    }
    Urgency = {0: 5, 1: 3, 2: 0, 3: 2}

    return {
        "status":         Status,
        "urgency":        Urgency.get(Label, 0),
        "reason":         Reasons.get(Status, ""),
        "soil_moisture":  Mois,
        "threshold_min":  T["Min"],
        "threshold_opt":  T["Opt"],
        "threshold_max":  T["Max"],
        "etc_mm":         round(ETC, 2),
        "deficit_mm":     round(Def, 2),
        "volume_l":       Vol,
        "recent_rain_mm": round(Rain3, 1),
        "heat_stress":    Weather_Today.get("평균온도_C", 0) > 33,
        "area_m2":        Area_m2,
        "label_int":      Label,
    }


# ══════════════════════════════════════════════════════════
#  섹션 G — AI 권고문 생성
# ══════════════════════════════════════════════════════════

def get_advice(
    Decision:      dict,
    Weather_Today: dict,
    Soil_Today:    dict,
    Crop_Type:     str,
    API_Key:       str | None = None,
) -> dict:
    """
    [INPUT]  의사결정 결과, 기상·토양 데이터, 작물, API 키
    [OUTPUT] {"advice": str, "source": "claude_ai" | "rule_based"}
    """
    import urllib.request, urllib.error

    Data = {
        "crop":         Crop_Type,
        "moisture":     Decision["soil_moisture"],
        "opt_moisture": Decision["threshold_opt"],
        "status":       Decision["status"],
        "temp":         Weather_Today.get("평균온도_C", 20),
        "humid":        Weather_Today.get("상대습도_Pct", 60),
        "rain":         Weather_Today.get("강수량_mm", 0),
        "rain_3day":    Decision["recent_rain_mm"],
        "et0":          Weather_Today.get("증발산량ET0_mm", 0),
        "deficit":      Decision["deficit_mm"],
        "volume":       Decision["volume_l"],
        "heat":         Decision["heat_stress"],
        "ph":           Soil_Today.get("pH", 6.5),
        "ec":           Soil_Today.get("전기전도도_dSm", 1.0),
    }

    if API_Key and _REQUESTS_OK:
        try:
            import json as _json
            Payload = _json.dumps({
                "model": "claude-sonnet-4-20250514", "max_tokens": 600,
                "system": "당신은 스마트 농업 전문 컨설턴트입니다. 주어진 기상·토양 데이터를 분석하여 농업인에게 실용적이고 구체적인 관수 권고문을 3~5문장으로 작성하세요. 수치를 반드시 포함하세요.",
                "messages": [{"role": "user", "content":
                    f"작물:{Data['crop']}, 토양수분:{Data['moisture']}%(적정:{Data['opt_moisture']}%), "
                    f"판단:{Data['status']}, 기온:{Data['temp']}°C, 습도:{Data['humid']}%, "
                    f"강수:{Data['rain']}mm(3일:{Data['rain_3day']}mm), ET0:{Data['et0']}mm, "
                    f"부족량:{Data['deficit']}mm, 권장관수:{Data['volume']}L, "
                    f"고온스트레스:{'있음' if Data['heat'] else '없음'}, "
                    f"pH:{Data['ph']}, EC:{Data['ec']}dS/m\n\n위 데이터로 관수 권고문 작성해주세요."
                }]
            }).encode("utf-8")
            Req = urllib.request.Request(
                "https://api.anthropic.com/v1/messages", data=Payload,
                headers={"Content-Type": "application/json",
                         "x-api-key": API_Key, "anthropic-version": "2023-06-01"},
                method="POST"
            )
            with urllib.request.urlopen(Req, timeout=15) as Resp:
                import json as _j2
                Txt = _j2.loads(Resp.read())["content"][0]["text"]
                return {"advice": Txt, "source": "claude_ai"}
        except Exception:
            pass

    # 규칙 기반 Fallback
    Lines = []
    S = Data["status"]
    if S == "즉시관수":
        Lines.append(f"⚠ {Data['crop']} 재배지의 토양 수분 {Data['moisture']}%가 심각하게 부족합니다. 즉시 약 {Data['volume']:.0f}L의 관수가 필요합니다.")
        Lines.append("이른 아침(06~08시) 또는 저녁(18~20시)에 점적관수 방식으로 공급하면 증발 손실을 최소화할 수 있습니다.")
        if Data["heat"]: Lines.append(f"기온 {Data['temp']}°C로 고온 스트레스 구간 — 잎 직접 살수는 화상 위험이 있으므로 지표 관수를 권장합니다.")
    elif S == "관수권장":
        Lines.append(f"{Data['crop']}의 토양 수분 {Data['moisture']}%는 적정보다 다소 낮습니다. 오늘 {Data['volume']:.0f}L 정도 보충을 권장합니다.")
        if Data["rain_3day"] > 10: Lines.append(f"최근 3일 강수 {data['rain_3day']}mm 감안하여 권장량의 60~70%만 관수해도 충분할 수 있습니다.")
    elif S == "적정상태":
        Lines.append(f"{Data['crop']} 재배지 토양 수분 {Data['moisture']}%는 현재 최적 범위입니다. 오늘 관수는 불필요합니다.")
        Lines.append("이틀 후 토양 수분을 재점검하고 기상 예보를 확인하세요.")
    else:  # 과습주의
        Lines.append(f"토양 수분 {Data['moisture']}%로 과습 상태입니다. 관수를 즉시 중단하고 배수로를 점검해 주세요.")
        Lines.append("과습이 지속되면 뿌리 호흡이 방해되어 생육 불량·병해 발생 위험이 높아집니다.")
    if Data["ph"] < 5.8: Lines.append(f"토양 pH {Data['ph']}로 산성 — 관수 시 석회 처리를 병행하세요.")
    if Data["ec"] > 2.0: Lines.append(f"EC {Data['ec']} dS/m로 염류 집적 가능성 — 충분한 관수로 용탈하세요.")

    return {"advice": "\n".join(Lines), "source": "rule_based"}


# ══════════════════════════════════════════════════════════
#  섹션 H — 이력 관리
# ══════════════════════════════════════════════════════════

def load_history(File_Path: str = "irrigation_history.json") -> list[dict]:
    """
    [INPUT]  File_Path
    [OUTPUT] 관수 이력 리스트
    """
    if not os.path.exists(File_Path): return []
    try:
        with open(File_Path, "r", encoding="utf-8") as F:
            Data = json.load(F)
            return Data if isinstance(Data, list) else []
    except Exception: return []


def save_history(Records: list[dict], File_Path: str = "irrigation_history.json") -> bool:
    """
    [INPUT]  Records 리스트, File_Path
    [OUTPUT] 성공 여부 (bool)
    """
    try:
        with open(File_Path, "w", encoding="utf-8") as F:
            json.dump(Records[-60:], F, ensure_ascii=False, indent=2)
        return True
    except Exception: return False


def add_history_record(
    Records:        list[dict],
    Date_Str:       str,
    Crop_Type:      str,
    Decision:       dict,
    Irrigated:      bool,
    Actual_Vol_L:   float = 0.0,
    File_Path:      str   = "irrigation_history.json",
) -> list[dict]:
    """
    [INPUT]  기존 이력 + 신규 레코드 데이터
    [OUTPUT] 업데이트된 이력 리스트
    """
    Records.append({
        "날짜": Date_Str, "작물": Crop_Type,
        "토양수분_Pct": Decision["soil_moisture"],
        "판단결과": Decision["status"],
        "긴급도": Decision["urgency"],
        "관수여부": Irrigated,
        "실제관수량_L": Actual_Vol_L,
        "권장관수량_L": Decision["volume_l"],
    })
    save_history(Records, File_Path)
    return Records


def calc_stats(Records: list[dict]) -> dict:
    """
    [INPUT]  이력 리스트
    [OUTPUT] {
        "total": int, "irrigated": int,
        "total_vol_L": float, "saved_vol_L": float,
        "status_dist": dict,
    }
    """
    Irr   = [R for R in Records if R.get("관수여부")]
    Dist  = {}
    for R in Records:
        S = R.get("판단결과", "알 수 없음")
        Dist[S] = Dist.get(S, 0) + 1
    Total_Vol = sum(R.get("실제관수량_L", 0) for R in Irr)
    Rec_Vol   = sum(R.get("권장관수량_L", 0) for R in Records)
    return {
        "total": len(Records), "irrigated": len(Irr),
        "total_vol_L": Total_Vol,
        "saved_vol_L": max(0, Rec_Vol - Total_Vol),
        "status_dist": Dist,
    }


# ══════════════════════════════════════════════════════════
#  섹션 I — 데이터 내보내기
# ══════════════════════════════════════════════════════════

def export_data(
    Weather_List:    list[dict],
    Soil_List:       list[dict],
    History_Records: list[dict],
    Base_Name:       str = "export",
    Save_Dir:        str = "data/processed",
) -> dict:
    """
    [INPUT]  기상·토양·이력 리스트, 파일 이름 기반, 저장 폴더
    [OUTPUT] {"csv_paths": [...], "xlsx_paths": [...], "log": str}

    CSV  — LLM 분석에 최적 (구분자 기반 텍스트)
    Excel — 수치 조정 · 시각화에 최적 (셀 서식 포함)
    """
    os.makedirs(Save_Dir, exist_ok=True)
    Ts        = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    CSV_Paths = []; XLSX_Paths = []; Logs = []

    Datasets = [
        (Weather_List,    f"{Base_Name}_weather_{Ts}"),
        (Soil_List,       f"{Base_Name}_soil_{Ts}"),
        (History_Records, f"{Base_Name}_history_{Ts}"),
    ]
    for Data, Name in Datasets:
        if not Data: continue
        Keys = list(Data[0].keys())
        # CSV
        try:
            CP = os.path.join(Save_Dir, f"{Name}.csv")
            with open(CP, "w", newline="", encoding="utf-8-sig") as F:
                W = csv.DictWriter(F, fieldnames=Keys, extrasaction="ignore")
                W.writeheader(); W.writerows(Data)
            CSV_Paths.append(CP)
        except Exception as E:
            Logs.append(f"CSV 오류 [{Name}]: {E}")
        # Excel
        if _EXCEL_OK:
            try:
                XP = os.path.join(Save_Dir, f"{Name}.xlsx")
                WB = openpyxl.Workbook(); WS = WB.active; WS.title = "데이터"
                WS.append(Keys)
                for R in Data: WS.append([R.get(K, "") for K in Keys])
                WB.save(XP); XLSX_Paths.append(XP)
            except Exception as E:
                Logs.append(f"Excel 오류 [{Name}]: {E}")

    return {
        "csv_paths":  CSV_Paths,
        "xlsx_paths": XLSX_Paths,
        "log":        f"CSV {len(CSV_Paths)}개, Excel {len(XLSX_Paths)}개 저장 완료" + ((" | 오류: " + ", ".join(Logs)) if Logs else ""),
    }